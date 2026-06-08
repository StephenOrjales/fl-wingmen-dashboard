import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from datetime import datetime, timedelta
import json
import re

st.set_page_config(
    page_title="FL Wingmen Dashboard",
    page_icon="\U0001F357",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .stApp { background-color: #FAFBFC; }
    .block-container { padding-top: 1rem; padding-bottom: 1rem; }

    .dash-header {
        background: linear-gradient(135deg, #1A3C34 0%, #2D6A4F 60%, #40916C 100%);
        padding: 1.1rem 1.8rem;
        border-radius: 12px;
        margin-bottom: 1.2rem;
        box-shadow: 0 4px 12px rgba(26,60,52,0.15);
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    .dash-header-left h1 {
        color: #FFFFFF;
        font-size: 1.3rem;
        font-weight: 800;
        margin: 0;
        letter-spacing: -0.3px;
    }
    .dash-header-left p {
        color: rgba(183,228,199,0.85);
        font-size: 0.72rem;
        margin: 0.2rem 0 0 0;
        font-weight: 500;
        letter-spacing: 0.3px;
    }
    .dash-header-right {
        display: flex;
        gap: 0.6rem;
        align-items: center;
    }
    .dash-chip {
        background: rgba(255,255,255,0.12);
        border: 1px solid rgba(255,255,255,0.15);
        border-radius: 20px;
        padding: 0.3rem 0.75rem;
        color: rgba(255,255,255,0.9);
        font-size: 0.7rem;
        font-weight: 500;
        white-space: nowrap;
    }

    .kpi-box {
        background: #FFFFFF;
        border: 1px solid #E8ECF0;
        border-radius: 10px;
        padding: 1rem 0.8rem;
        text-align: center;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .kpi-label {
        color: #6B7280;
        font-size: 0.68rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.8px;
    }
    .kpi-value {
        color: #1F2937;
        font-size: 1.5rem;
        font-weight: 700;
        margin-top: 0.2rem;
    }
    .kpi-value.green { color: #059669; }
    .kpi-value.orange { color: #D97706; }
    .kpi-value.red { color: #DC2626; }

    .section-title {
        color: #1F2937;
        font-size: 1rem;
        font-weight: 600;
        margin: 1rem 0 0.5rem 0;
        padding-bottom: 0.3rem;
        border-bottom: 2px solid #2D6A4F;
        display: inline-block;
    }

    div[data-testid="stMetric"] {
        background: #FFFFFF;
        border: 1px solid #E8ECF0;
        border-radius: 10px;
        padding: 0.8rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    div[data-testid="stMetric"] label { color: #6B7280 !important; font-size: 0.72rem !important; text-transform: uppercase; letter-spacing: 0.5px; }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #1F2937 !important; font-size: 1.4rem !important; font-weight: 700 !important; }

    .stTabs [data-baseweb="tab-list"] { display: none; }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FFFFFF 0%, #F8FAF9 100%);
        border-right: 1px solid #E2E8F0;
        width: 240px !important; min-width: 240px !important;
    }
    section[data-testid="stSidebar"] > div { width: 240px !important; padding-top: 0.5rem; }
    section[data-testid="stSidebar"] label { color: #374151 !important; font-size: 0.8rem !important; }
    section[data-testid="stSidebar"] .stMarkdown p { color: #6B7280; font-size: 0.8rem; }
    section[data-testid="stSidebar"] .stMarkdown hr { border: none; height: 1px; background: linear-gradient(90deg, transparent, #D1D5DB, transparent); margin: 0.6rem 0; }
    section[data-testid="stSidebar"] .stSelectbox > div > div {
        font-size: 0.82rem;
        border-radius: 8px;
        border: 1px solid #D1D5DB;
        background: #FFFFFF;
    }
    section[data-testid="stSidebar"] .stSelectbox > div > div:hover {
        border-color: #2D6A4F;
    }
    section[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] {
        border: none !important;
        background: transparent !important;
        box-shadow: none !important;
    }

    .stDataFrame { border-radius: 8px; overflow: hidden; }
    .stDataFrame [data-testid="stDataFrameResizable"] { background: #FFFFFF; border: 1px solid #E8ECF0; border-radius: 8px; }
    .stDataFrame th { background: #F0FDF4 !important; color: #1F2937 !important; font-size: 0.78rem !important; font-weight: 600 !important; border-bottom: 2px solid #2D6A4F !important; }
    .stDataFrame td { color: #374151 !important; font-size: 0.78rem !important; background: #FFFFFF !important; }
    .stDataFrame tr:hover td { background: #F9FAFB !important; }
    [data-testid="glideDataEditor"] { border: 1px solid #E8ECF0 !important; border-radius: 8px !important; }
    [data-testid="glideDataEditor"] th, [data-testid="glideDataEditor"] .header-cell { background: #F0FDF4 !important; color: #1F2937 !important; }
    [data-testid="stDataFrame"] > div { background: #FFFFFF; border-radius: 8px; }

    div[data-testid="stExpander"] { background: #FFFFFF; border: 1px solid #E8ECF0; border-radius: 8px; }
    div[data-testid="stExpander"] summary span { color: #374151 !important; }

    section[data-testid="stSidebar"] div[data-testid="stRadio"] > div { gap: 2px; }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label {
        background: transparent;
        color: #4B5563 !important;
        padding: 0.4rem 0.75rem;
        border-radius: 8px;
        font-weight: 500;
        font-size: 0.82rem;
        cursor: pointer;
        transition: all 0.2s ease;
        border: 1px solid transparent;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label p {
        color: #4B5563 !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:hover {
        background: #F0FDF4;
        color: #2D6A4F !important;
        border-color: #D1FAE5;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:hover p {
        color: #2D6A4F !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label[data-checked="true"],
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:has(input:checked) {
        background: linear-gradient(135deg, #1A3C34 0%, #2D6A4F 100%);
        color: #FFFFFF !important;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(26,60,52,0.2);
        border-color: transparent;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label[data-checked="true"] p,
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:has(input:checked) p {
        color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label span p { margin: 0; }
</style>
""", unsafe_allow_html=True)

DATA_DIR = Path(__file__).parent / "data"
CONFIG_PATH = Path(__file__).parent / "config.json"

DISTRICTS = {}
if CONFIG_PATH.exists():
    with open(CONFIG_PATH) as f:
        config = json.load(f)
        DISTRICTS = config.get("stores", {})

STORE_TO_DISTRICT = {}
for district, stores in DISTRICTS.items():
    for store in stores:
        store_num = store.split(" - ")[0].strip().lstrip("0")
        STORE_TO_DISTRICT[store_num] = district

GREEN = "#059669"
GOLD = "#D97706"
TEAL = "#0D9488"
ORANGE = "#D97706"
RED = "#DC2626"
DARK = "#64748B"
CHART_BG = "#FFFFFF"
GRID_COLOR = "#F1F5F9"
FONT_COLOR = "#374151"

CHART_LAYOUT = dict(
    plot_bgcolor=CHART_BG,
    paper_bgcolor=CHART_BG,
    font=dict(color=FONT_COLOR, size=11),
    margin=dict(l=50, r=20, t=30, b=90),
    xaxis=dict(gridcolor=GRID_COLOR, tickfont=dict(size=9), fixedrange=True),
    yaxis=dict(gridcolor=GRID_COLOR, fixedrange=True),
    dragmode=False,
)

CHART_CONFIG = dict(
    displayModeBar=False,
    scrollZoom=False,
)


def fiscal_week_label(d):
    """Wingstop fiscal weeks run Sunday-Saturday. Returns (week_start_date, label)."""
    from datetime import date
    if not isinstance(d, date):
        return None, "Unknown"
    days_since_sun = (d.weekday() + 1) % 7
    week_start = d - timedelta(days=days_since_sun)
    fy_start = date(2025, 12, 28)
    week_num = ((week_start - fy_start).days // 7) + 1
    return week_start, f"W{week_num}"


def fmt_sos(minutes):
    """Format decimal minutes (e.g. 8.5) as M:SS (e.g. 8:30)."""
    if pd.isna(minutes):
        return "-"
    m = int(minutes)
    s = int(round((minutes - m) * 60))
    return f"{m}:{s:02d}"


def parse_time_to_minutes(t):
    if pd.isna(t) or not isinstance(t, str):
        return None
    m = re.match(r"(\d+):(\d+)", str(t))
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60
    return None


def parse_pct(v):
    if pd.isna(v) or not isinstance(v, str):
        return None
    v = str(v).replace("%", "").strip()
    try:
        return float(v)
    except ValueError:
        return None


def extract_store_number(name):
    m = re.match(r"(\d+)", str(name))
    return m.group(1) if m else None


def short_name(store):
    parts = str(store).split("-", 2)
    if len(parts) >= 3:
        return parts[2].strip()[:22]
    return str(store)[:22]


def forecast_short_name(store):
    parts = str(store).split(" - ", 1)
    if len(parts) == 2:
        return parts[1].strip()[:22]
    return str(store)[:22]


def forecast_store_num(store):
    m = re.match(r"0*(\d+)", str(store))
    return m.group(1) if m else None



# KDS data removed — will be replaced with new Fri/Sat Dinner adherence system



def _load_forecast_quarter(periods):
    path = DATA_DIR / "forecast.xlsm"
    if not path.exists():
        return pd.DataFrame(), pd.DataFrame()
    raw = pd.read_excel(path, sheet_name="Forecast_Data")
    qtr = raw[(raw["year"] == 2026) & (raw["period"].isin(periods))].copy()
    if qtr.empty:
        return pd.DataFrame(), pd.DataFrame()

    store_agg = qtr.groupby(["store", "district"]).agg(
        forecast_sales=("forecast_sales", "sum"),
        actual_sales=("actual_sales", "sum"),
        guide_hours=("guide_hours", "sum"),
        schedule_hours=("schedule_hours", "sum"),
        actual_hours=("actual_crew_hours", "sum"),
        schedule_labor=("schedule_labor", "sum"),
        actual_labor=("actual_labor", "sum"),
        ovt_hours=("ovt_hours", "sum"),
        weeks=("week_d", "count"),
    ).reset_index()

    store_agg["sales_var_pct"] = (store_agg["actual_sales"] - store_agg["forecast_sales"]) / store_agg["forecast_sales"]
    store_agg["labor_var_pct"] = (store_agg["actual_hours"] - store_agg["guide_hours"]) / store_agg["guide_hours"]
    store_agg["actual_labor_pct"] = store_agg["actual_labor"] / store_agg["actual_sales"]
    store_agg["schedule_labor_pct"] = store_agg["schedule_labor"] / store_agg["forecast_sales"]
    store_agg["labor_pct_variance"] = store_agg["actual_labor_pct"] - store_agg["schedule_labor_pct"]
    store_agg["store_num"] = store_agg["store"].apply(forecast_store_num)
    store_agg["short_name"] = store_agg["store"].apply(forecast_short_name)
    store_agg["config_district"] = store_agg["store_num"].map(STORE_TO_DISTRICT)

    qtr["store_num"] = qtr["store"].apply(forecast_store_num)
    qtr["config_district"] = qtr["store_num"].map(STORE_TO_DISTRICT)
    qtr["actual_labor_pct"] = qtr["actual_labor"] / qtr["actual_sales"]
    qtr["schedule_labor_pct"] = qtr["schedule_labor"] / qtr["forecast_sales"]
    qtr["labor_pct_variance"] = qtr["actual_labor_pct"] - qtr["schedule_labor_pct"]
    qtr["short_name"] = qtr["store"].apply(forecast_short_name)

    return store_agg, qtr


@st.cache_data(ttl=300)
def load_q1_data():
    return _load_forecast_quarter([1, 2, 3])


@st.cache_data(ttl=300)
def load_q2_data():
    return _load_forecast_quarter([4, 5, 6])


kds_df_all = pd.DataFrame()
daily_df_all = pd.DataFrame()
kds_q1_all = pd.DataFrame()
q1_store, q1_weekly = load_q1_data()
q2_store, q2_weekly = load_q2_data()

# ── Sidebar ──
with st.sidebar:
    # ── Logo / Brand ──
    st.markdown("""
    <div style="text-align:center; padding:0.6rem 0 0.2rem 0;">
        <div style="font-size:1.3rem; font-weight:800; color:#1A3C34; letter-spacing:-0.5px;">FL Wingmen</div>
        <div style="font-size:0.65rem; color:#6B7280; text-transform:uppercase; letter-spacing:1.5px; margin-top:2px;">Operations Dashboard</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    # ── Filters ──
    st.markdown("""
    <div style="display:flex; align-items:center; gap:6px; margin-bottom:0.5rem;">
        <div style="width:3px; height:16px; background:#2D6A4F; border-radius:2px;"></div>
        <span style="color:#374151; font-weight:600; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px;">Filters</span>
    </div>
    """, unsafe_allow_html=True)

    district_options = ["All Districts"] + sorted(DISTRICTS.keys())
    selected_district = st.selectbox("District", district_options, label_visibility="collapsed")

    # Build store list from config
    all_config_stores = []
    for dist, stores in DISTRICTS.items():
        for s in stores:
            snum = s.split(" - ")[0].strip().lstrip("0")
            all_config_stores.append({"Store Full Name": s, "store_num": snum, "district": dist})
    store_src = pd.DataFrame(all_config_stores) if all_config_stores else pd.DataFrame(columns=["Store Full Name", "store_num"])

    if selected_district == "All Districts":
        store_list = sorted(store_src["Store Full Name"].dropna().unique())
    else:
        district_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
        store_list = sorted(store_src[store_src["store_num"].isin(district_nums)]["Store Full Name"].dropna().unique())

    store_options = ["All Stores"] + store_list
    selected_store = st.selectbox("Store", store_options, label_visibility="collapsed")

    st.markdown("---")

    # ── Navigation ──
    st.markdown("""
    <div style="display:flex; align-items:center; gap:6px; margin-bottom:0.3rem;">
        <div style="width:3px; height:16px; background:#2D6A4F; border-radius:2px;"></div>
        <span style="color:#374151; font-weight:600; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px;">Navigation</span>
    </div>
    """, unsafe_allow_html=True)

    nav_options = ["Sales Performance", "KDS Dashboard", "Schedule Guide", "Internal QSC Evals", "Labor Dashboard", "COGS Variance", "SMG (Guest Satisfaction)", "District Comparison", "Scorecard", "Watch List", "Wing Worm"]
    selected_tab = st.radio("Nav", nav_options, label_visibility="collapsed")

    st.markdown("---")
    st.markdown(f"""
    <div style="text-align:center; padding:0.2rem 0;">
        <span style="color:#9CA3AF; font-size:0.7rem;">{len(store_src)} stores &nbsp;&bull;&nbsp; {len(DISTRICTS)} districts</span>
    </div>
    """, unsafe_allow_html=True)

# ── Header ──
filter_text = "All Stores"
if selected_store != "All Stores":
    filter_text = short_name(selected_store)
elif selected_district != "All Districts":
    filter_text = selected_district

_today_str = datetime.now().strftime("%b %d, %Y")
st.markdown(f"""
<div class="dash-header">
    <div class="dash-header-left">
        <h1>FL Wingmen Dashboard</h1>
        <p>{filter_text}</p>
    </div>
    <div class="dash-header-right">
        <span class="dash-chip">{len(DISTRICTS)} Districts</span>
        <span class="dash-chip">{len(store_src)} Stores</span>
        <span class="dash-chip">{_today_str}</span>
    </div>
</div>
""", unsafe_allow_html=True)


def kpi_card(label, value, color=""):
    cls = f" {color}" if color else ""
    return f"""<div class="kpi-box">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value{cls}">{value}</div>
    </div>"""


# ════════════════════════════════
# KDS DASHBOARD (new polished version)
# ════════════════════════════════
if selected_tab == "KDS Dashboard":

    kds_file = DATA_DIR / "kds_dinner.csv"
    if kds_file.exists():
        kds_raw = pd.read_csv(kds_file)
        periods_sorted = sorted(kds_raw["Period"].unique(), key=lambda x: (int(x[1]), int(x[3])))

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = extract_store_number(selected_store).lstrip("0")
            kds_raw = kds_raw[kds_raw["Store No"].astype(str) == sk_num]
        elif selected_district != "All Districts":
            d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
            kds_raw = kds_raw[kds_raw["Store No"].astype(str).isin(d_nums)]

        periods_sorted = [p for p in periods_sorted if p in kds_raw["Period"].unique()]
        latest_period = periods_sorted[-1] if periods_sorted else ""

        # ── HEADER ──
        st.markdown(f"""
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
            <div>
                <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">SOS Adherence Dashboard</h2>
                <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                    FL Wingmen — {len(DISTRICTS)} districts, {len(STORE_TO_DISTRICT)} stores
                </p>
            </div>
            <div style="display:flex; gap:0.5rem; align-items:center;">
                <div style="background:#D97706; color:#FFFFFF; padding:0.5rem 1rem; border-radius:8px; font-weight:700; font-size:0.8rem; white-space:nowrap;">
                    FRI &amp; SAT DINNER ONLY
                </div>
                <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:8px; font-weight:700; font-size:0.8rem; white-space:nowrap;">
                    TARGET · <span style="color:#FFD700;">SOS &lt; 10 min</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Period selector + Export ──
        pcol1, pcol2, pcol3 = st.columns([1, 1, 2])
        with pcol1:
            period_options = list(reversed(periods_sorted))
            sel_period = st.selectbox("Week", period_options, index=0, key="kds_dash_period", label_visibility="collapsed")

        kds_week = kds_raw[kds_raw["Period"] == sel_period].copy()
        kds_week["District"] = kds_week["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("Unassigned")

        # SOS classification
        def classify_sos(row):
            if pd.isna(row["SOS"]):
                return "No Data"
            if row["SOS"] < 7:
                return "Fast"
            if row["SOS"] <= 10:
                return "Adherent"
            return "Slow"

        kds_week["SOS Status"] = kds_week.apply(classify_sos, axis=1)

        # Overall adherence classification
        def classify_adherence(row):
            if pd.isna(row["Adherence %"]):
                return "No Data"
            if row["Adherence %"] >= 80:
                return "Adherent"
            if row["Adherence %"] >= 60:
                return "Needs Improvement"
            return "Non-Adherent"

        kds_week["Overall Status"] = kds_week.apply(classify_adherence, axis=1)

        n_total_config = len(STORE_TO_DISTRICT)
        n_reporting = kds_week["Store No"].nunique()
        n_adherent = (kds_week["SOS Status"] == "Adherent").sum()
        n_slow = (kds_week["SOS Status"] == "Slow").sum()
        n_fast = (kds_week["SOS Status"] == "Fast").sum()
        n_nodata = (kds_week["SOS Status"] == "No Data").sum()
        pct_adherent = (n_adherent / n_reporting * 100) if n_reporting > 0 else 0

        # ── Excel Export (styled) ──
        import io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def style_excel_sheet(ws, df, pct_cols=None, dec1_cols=None, dec2_cols=None, status_col=None):
            """Apply consistent styling to an Excel worksheet."""
            header_fill = PatternFill(start_color="1A3C34", end_color="1A3C34", fill_type="solid")
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            cell_font = Font(name="Calibri", size=10)
            thin_border = Border(
                bottom=Side(style="thin", color="E2E8F0"),
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
            )
            header_border = Border(
                bottom=Side(style="medium", color="1A3C34"),
                left=Side(style="thin", color="1A3C34"),
                right=Side(style="thin", color="1A3C34"),
            )
            green_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            status_fills = {"Adherent": green_fill, "Slow": red_fill, "Fast": yellow_fill, "No Data": gray_fill}

            pct_cols = pct_cols or []
            dec1_cols = dec1_cols or []
            dec2_cols = dec2_cols or []

            # Style header row
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_border

            # Style data rows
            status_col_idx = list(df.columns).index(status_col) + 1 if status_col and status_col in df.columns else None
            for row_idx in range(2, len(df) + 2):
                # Determine row fill from status
                row_fill = None
                if status_col_idx:
                    status_val = ws.cell(row=row_idx, column=status_col_idx).value
                    row_fill = status_fills.get(status_val)

                # Alternate row shading if no status coloring
                if not row_fill and row_idx % 2 == 0:
                    row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if row_fill:
                        cell.fill = row_fill

                    # Number formatting
                    col_name = df.columns[col_idx - 1]
                    if col_name in pct_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = '0"%"'
                    elif col_name in dec1_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'
                    elif col_name in dec2_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            # Auto-fit column widths
            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = len(str(col_name)) + 2
                for row_idx in range(2, min(len(df) + 2, 52)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 1, 35)

            # Freeze top row
            ws.freeze_panes = "A2"

        with pcol2:
            st.markdown("<div style='height:0.3rem;'></div>", unsafe_allow_html=True)
            export_buf = io.BytesIO()
            with pd.ExcelWriter(export_buf, engine="openpyxl") as writer:
                # Sheet 1: Current week detail
                exp_week = kds_week[["Store No", "Store Name", "District", "SOS", "SOS Status", "Overall Status",
                                     "Adoption %", "Make Ahead %", "Waste %", "Pre-Bump %", "Adherence %"]].copy()
                exp_week = exp_week.sort_values("SOS", na_position="last")
                exp_week.to_excel(writer, sheet_name=sel_period, index=False)
                style_excel_sheet(writer.sheets[sel_period], exp_week,
                                  pct_cols=["Adherence %"], dec1_cols=["SOS", "Adoption %", "Make Ahead %"],
                                  dec2_cols=["Waste %", "Pre-Bump %"], status_col="SOS Status")

                # Sheet 2: All periods raw (trimmed columns)
                exp_all = kds_raw.copy()
                exp_all["District"] = exp_all["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("Unassigned")
                exp_all["SOS Status"] = exp_all.apply(classify_sos, axis=1)
                exp_all = exp_all[["Period", "Store No", "Store Name", "District", "SOS", "SOS Status",
                                   "Adoption %", "Make Ahead %", "Waste %", "Pre-Bump %", "Adherence %"]].copy()
                exp_all = exp_all.sort_values(["Period", "Store No"])
                exp_all.to_excel(writer, sheet_name="All Periods", index=False)
                style_excel_sheet(writer.sheets["All Periods"], exp_all,
                                  pct_cols=["Adherence %"], dec1_cols=["SOS", "Adoption %", "Make Ahead %"],
                                  dec2_cols=["Waste %", "Pre-Bump %"], status_col="SOS Status")

                # Sheet 3: Historical summary
                hist_export = []
                for p in periods_sorted:
                    pw = kds_raw[kds_raw["Period"] == p]
                    pw_valid = pw[pw["SOS"].notna()]
                    n_pw = len(pw_valid)
                    n_adh_p = pw_valid["SOS"].between(0, 10, inclusive="right").sum()
                    hist_export.append({
                        "Period": p, "Stores": n_pw,
                        "Adherent": n_adh_p, "Slow": (pw_valid["SOS"] > 10).sum(), "Fast": (pw_valid["SOS"] < 7).sum(),
                        "Adherent %": round(n_adh_p / n_pw * 100, 1) if n_pw else 0,
                        "Avg SOS": round(pw_valid["SOS"].mean(), 1) if len(pw_valid) else 0,
                        "Avg Adherence %": round(pw["Adherence %"].mean(), 0) if pw["Adherence %"].notna().any() else 0,
                    })
                hist_df_exp = pd.DataFrame(hist_export)
                hist_df_exp.to_excel(writer, sheet_name="Summary", index=False)
                style_excel_sheet(writer.sheets["Summary"], hist_df_exp,
                                  dec1_cols=["Adherent %", "Avg SOS", "Avg Adherence %"])
            export_buf.seek(0)
            st.download_button("📥 Export Excel", data=export_buf, file_name=f"KDS_Dashboard_{sel_period}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="kds_export")

        # ── SUB-TABS ──
        tab_overview, tab_by_district, tab_offenders, tab_historical = st.tabs(
            [f"📊 {sel_period} Overview", "🏢 By District", "🔴 Repeat Offenders", f"📈 Historical ({periods_sorted[0]}–{periods_sorted[-1]})"]
        )

        # ════════════════════
        # TAB: OVERVIEW
        # ════════════════════
        with tab_overview:
            # ── KPI Cards ──
            c1, c2, c3, c4, c5 = st.columns(5)
            kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
                <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
                <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
                <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
            </div>"""

            c1.markdown(kpi_style.format(label="REPORTING STORES", value=n_reporting, color="#1F2937",
                        sub=f"of {n_total_config} total"), unsafe_allow_html=True)
            c2.markdown(kpi_style.format(label="ADHERENT (SOS 7-10M)", value=n_adherent, color="#059669",
                        sub=f"{pct_adherent:.1f}% of reporting"), unsafe_allow_html=True)
            c3.markdown(kpi_style.format(label="SLOW OUTLIERS", value=n_slow, color="#DC2626",
                        sub=f"{(n_slow/n_reporting*100) if n_reporting else 0:.0f}% of reporting"), unsafe_allow_html=True)
            c4.markdown(kpi_style.format(label="FAST OUTLIERS", value=n_fast, color="#D97706",
                        sub=f"{(n_fast/n_reporting*100) if n_reporting else 0:.0f}% of reporting"), unsafe_allow_html=True)
            avg_sos_kpi = kds_week["SOS"].mean() if kds_week["SOS"].notna().any() else 0
            sos_kpi_color = "#059669" if avg_sos_kpi < 10 else ("#D97706" if avg_sos_kpi < 13 else "#DC2626")
            c5.markdown(kpi_style.format(label="AVG SOS", value=fmt_sos(avg_sos_kpi), color=sos_kpi_color,
                        sub="target < 10 min"), unsafe_allow_html=True)

            st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

            # ── Charts row: Donut + District Distribution ──
            chart_l, chart_r = st.columns(2)

            with chart_l:
                st.markdown(f"""<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem;">
                    <div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">
                        <span style="background:#1A3C34; color:#FFFFFF; padding:2px 10px; border-radius:4px; font-size:0.75rem; font-weight:700;">{sel_period}</span>
                        <span style="font-weight:700; color:#1F2937;">Adherence Mix</span>
                    </div>
                </div>""", unsafe_allow_html=True)

                donut_data = kds_week["SOS Status"].value_counts()
                color_map = {"Adherent": "#059669", "Slow": "#DC2626", "Fast": "#D97706", "No Data": "#CBD5E1"}
                labels = donut_data.index.tolist()
                values = donut_data.values.tolist()
                colors = [color_map.get(l, "#CBD5E1") for l in labels]
                legend_labels = [f"Adherent (7-10m)" if l == "Adherent" else f"Slow Outlier (>10m)" if l == "Slow" else f"Fast Outlier (<7m)" if l == "Fast" else "No Data" for l in labels]

                fig_donut = go.Figure(go.Pie(
                    labels=legend_labels, values=values, hole=0.55,
                    marker=dict(colors=colors),
                    textinfo="percent", textfont=dict(size=13, color="#FFFFFF"),
                    hovertemplate="%{label}: %{value} stores (%{percent})<extra></extra>",
                    sort=False,
                ))
                fig_donut.update_layout(
                    plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
                    font=dict(color=FONT_COLOR, size=11),
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=320,
                    legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05, font=dict(size=11)),
                    showlegend=True,
                )
                st.plotly_chart(fig_donut, use_container_width=True, config=CHART_CONFIG)

            with chart_r:
                st.markdown(f"""<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem;">
                    <div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">
                        <span style="background:#1A3C34; color:#FFFFFF; padding:2px 10px; border-radius:4px; font-size:0.75rem; font-weight:700;">{sel_period}</span>
                        <span style="font-weight:700; color:#1F2937;">Distribution by District</span>
                    </div>
                </div>""", unsafe_allow_html=True)

                dist_status = kds_week.groupby(["District", "SOS Status"]).size().unstack(fill_value=0)
                for col in ["Adherent", "Slow", "Fast", "No Data"]:
                    if col not in dist_status.columns:
                        dist_status[col] = 0

                fig_dist = go.Figure()
                for status, color in [("Adherent", "#059669"), ("Slow", "#DC2626"), ("Fast", "#D97706"), ("No Data", "#CBD5E1")]:
                    fig_dist.add_trace(go.Bar(
                        x=dist_status.index, y=dist_status[status], name=status,
                        marker_color=color, hovertemplate="%{x}<br>" + status + ": %{y}<extra></extra>",
                    ))
                dist_bar_layout = {**CHART_LAYOUT, "barmode": "stack",
                                   "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category", tickfont=dict(size=9)),
                                   "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, dtick=5),
                                   "legend": dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5, font=dict(size=10)),
                                   "margin": dict(l=40, r=10, t=10, b=60)}
                fig_dist.update_layout(**dist_bar_layout, height=320)
                st.plotly_chart(fig_dist, use_container_width=True, config=CHART_CONFIG)

            # ── Takeaways ──
            st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.05rem; margin:1rem 0 0.5rem 0;">{sel_period} Takeaways</div>', unsafe_allow_html=True)

            # Auto-generate insights
            takeaway_style = '<div style="border-left:4px solid {color}; padding:0.5rem 1rem; margin:0.4rem 0; background:#FAFBFC; border-radius:0 6px 6px 0;">{text}</div>'

            # 1. Overall adherence
            prev_period = periods_sorted[periods_sorted.index(sel_period) - 1] if periods_sorted.index(sel_period) > 0 else None
            delta_text = ""
            if prev_period:
                prev_week = kds_raw[kds_raw["Period"] == prev_period]
                prev_pct = (prev_week["SOS"].dropna().between(0, 10, inclusive="right").sum() / len(prev_week) * 100) if len(prev_week) > 0 else 0
                curr_pct = pct_adherent
                delta = curr_pct - prev_pct
                arrow = "▲" if delta > 0 else "▼"
                delta_color = "#059669" if delta > 0 else "#DC2626"
                delta_text = f' <span style="color:{delta_color}; font-weight:700;">{arrow} {abs(delta):.1f} pp vs {prev_period}</span>.'

            st.markdown(takeaway_style.format(color="#1A3C34",
                text=f'<b>{pct_adherent:.1f}% adherent</b> across {n_reporting} reporting stores.{delta_text} {n_slow} slow, {n_fast} fast, {n_nodata} no data.'),
                unsafe_allow_html=True)

            # 2. Best & worst district
            dist_adh_rates = kds_week.groupby("District").apply(
                lambda g: (g["SOS"].dropna().between(0, 10, inclusive="right").sum() / len(g) * 100) if len(g) > 0 else 0
            ).sort_values(ascending=False)
            if len(dist_adh_rates) > 0:
                best_dist = dist_adh_rates.index[0]
                best_dist_n = kds_week[kds_week["District"] == best_dist]["SOS"].dropna().between(0, 10, inclusive="right").sum()
                best_dist_total = len(kds_week[kds_week["District"] == best_dist])
                st.markdown(takeaway_style.format(color="#059669",
                    text=f'<b>Best District:</b> {best_dist} — {int(best_dist_n)}/{best_dist_total} adherent ({dist_adh_rates.iloc[0]:.0f}%).'),
                    unsafe_allow_html=True)
            if len(dist_adh_rates) > 1:
                worst_dist = dist_adh_rates.index[-1]
                worst_dist_n = kds_week[kds_week["District"] == worst_dist]["SOS"].dropna().between(0, 10, inclusive="right").sum()
                worst_dist_total = len(kds_week[kds_week["District"] == worst_dist])
                st.markdown(takeaway_style.format(color="#DC2626",
                    text=f'<b>Worst District:</b> {worst_dist} — {int(worst_dist_n)}/{worst_dist_total} adherent ({dist_adh_rates.iloc[-1]:.0f}%).'),
                    unsafe_allow_html=True)

            # 3. Worst & best single store
            sos_valid = kds_week[kds_week["SOS"].notna()].copy()
            if len(sos_valid) > 0:
                worst_store = sos_valid.loc[sos_valid["SOS"].idxmax()]
                best_store = sos_valid.loc[sos_valid["SOS"].idxmin()]
                st.markdown(takeaway_style.format(color="#DC2626",
                    text=f'<b>Worst single store:</b> {int(worst_store["Store No"])} {worst_store["Store Name"]} ({fmt_sos(worst_store["SOS"])}). '
                         f'<b>Fastest:</b> {int(best_store["Store No"])} {best_store["Store Name"]} ({fmt_sos(best_store["SOS"])}).'),
                    unsafe_allow_html=True)

            st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

            # ── Store Performance Table ──
            st.markdown(f"""<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">
                <span style="background:#1A3C34; color:#FFFFFF; padding:2px 10px; border-radius:4px; font-size:0.75rem; font-weight:700;">{sel_period}</span>
                <span style="font-weight:700; color:#1F2937; font-size:1.05rem;">Store Performance</span>
                <span style="color:#9CA3AF; font-size:0.75rem; margin-left:0.5rem;">Click any column header to sort</span>
            </div>""", unsafe_allow_html=True)

            DISTRICT_COLORS = {
                "District 1": "#EC4899", "District 2": "#7C3AED", "District 3": "#0369A1",
                "District 4": "#CA8A04", "District 5": "#059669", "District 6": "#4338CA",
            }
            OFF_GUIDE = "color: #DC2626; font-weight: 700"

            tbl = kds_week[["Store No", "Store Name", "District", "SOS", "SOS Status", "Adoption %", "Make Ahead %", "Waste %", "Pre-Bump %", "Adherence %"]].copy()
            tbl = tbl.sort_values("Store No", ascending=True)

            # Keep raw values for conditional styling
            raw_sos = tbl["SOS"].copy()
            raw_adopt = tbl["Adoption %"].copy()
            raw_ma = tbl["Make Ahead %"].copy()
            raw_waste = tbl["Waste %"].copy()
            raw_pb = tbl["Pre-Bump %"].copy()
            raw_adh = tbl["Adherence %"].copy()
            raw_district = tbl["District"].copy()

            tbl["SOS"] = tbl["SOS"].apply(fmt_sos)
            tbl["Adoption %"] = tbl["Adoption %"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
            tbl["Make Ahead %"] = tbl["Make Ahead %"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
            tbl["Waste %"] = tbl["Waste %"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "-")
            tbl["Pre-Bump %"] = tbl["Pre-Bump %"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "-")
            tbl["Adherence %"] = tbl["Adherence %"].apply(lambda x: f"{x:.0f}%" if pd.notna(x) else "-")
            tbl = tbl.rename(columns={"SOS Status": "Status", "Store No": "Store #"})
            tbl = tbl.reset_index(drop=True)
            raw_sos = raw_sos.reset_index(drop=True)
            raw_adopt = raw_adopt.reset_index(drop=True)
            raw_ma = raw_ma.reset_index(drop=True)
            raw_waste = raw_waste.reset_index(drop=True)
            raw_pb = raw_pb.reset_index(drop=True)
            raw_adh = raw_adh.reset_index(drop=True)
            raw_district = raw_district.reset_index(drop=True)

            def style_store_table(row):
                idx = row.name
                styles = [""] * len(row)
                cols = list(row.index)
                # District colored cell
                dist = raw_district.get(idx, "")
                d_color = DISTRICT_COLORS.get(dist, "#374151")
                styles[cols.index("District")] = f"background-color: {d_color}; color: #FFFFFF; font-weight: 600"
                # Off-guide highlights
                if pd.notna(raw_sos.get(idx)) and raw_sos[idx] > 10:
                    styles[cols.index("SOS")] += f"; {OFF_GUIDE}"
                if pd.notna(raw_adopt.get(idx)) and raw_adopt[idx] < 85:
                    styles[cols.index("Adoption %")] += f"; {OFF_GUIDE}"
                if pd.notna(raw_ma.get(idx)) and raw_ma[idx] > 0:
                    styles[cols.index("Make Ahead %")] += f"; {OFF_GUIDE}"
                if pd.notna(raw_waste.get(idx)) and raw_waste[idx] > 0:
                    styles[cols.index("Waste %")] += f"; {OFF_GUIDE}"
                if pd.notna(raw_pb.get(idx)) and raw_pb[idx] > 0.5:
                    styles[cols.index("Pre-Bump %")] += f"; {OFF_GUIDE}"
                if pd.notna(raw_adh.get(idx)) and raw_adh[idx] < 85:
                    styles[cols.index("Adherence %")] += f"; {OFF_GUIDE}"
                return styles

            styled_tbl = tbl.style.apply(style_store_table, axis=1)
            st.dataframe(styled_tbl, use_container_width=True, hide_index=True, height=500)

        # ════════════════════
        # TAB: BY DISTRICT
        # ════════════════════
        with tab_by_district:
            st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.1rem; margin-bottom:0.5rem;">District Breakdown — {sel_period}</div>', unsafe_allow_html=True)

            for district in sorted(kds_week["District"].unique()):
                d_data = kds_week[kds_week["District"] == district].copy()
                d_adherent = (d_data["SOS Status"] == "Adherent").sum()
                d_slow = (d_data["SOS Status"] == "Slow").sum()
                d_fast = (d_data["SOS Status"] == "Fast").sum()
                d_total = len(d_data)
                d_avg_sos = d_data["SOS"].mean() if d_data["SOS"].notna().any() else 0
                d_avg_adh = d_data["Adherence %"].mean() if d_data["Adherence %"].notna().any() else 0

                adh_pct = d_adherent / d_total * 100 if d_total > 0 else 0
                header_color = "#059669" if adh_pct >= 70 else ("#D97706" if adh_pct >= 50 else "#DC2626")

                badges = ""
                if d_slow > 0:
                    badges += f'<span style="background:#DC2626; color:#FFF; padding:2px 8px; border-radius:4px; font-size:0.75rem; margin-left:6px;">{d_slow} slow</span>'
                if d_fast > 0:
                    badges += f'<span style="background:#D97706; color:#FFF; padding:2px 8px; border-radius:4px; font-size:0.75rem; margin-left:4px;">{d_fast} fast</span>'

                st.markdown(f"""
                <div style="background:#1A3C34; color:#FFFFFF; padding:0.6rem 1rem; border-radius:8px 8px 0 0; margin-top:1rem;
                            display:flex; justify-content:space-between; align-items:center;">
                    <span style="font-weight:700; font-size:0.95rem;">{district}{badges}</span>
                    <span style="font-size:0.82rem;">Adherent: <b>{d_adherent}/{d_total}</b> ({adh_pct:.0f}%) &nbsp;|&nbsp; Avg SOS: <b>{fmt_sos(d_avg_sos)}</b> &nbsp;|&nbsp; Avg Adherence: <b>{d_avg_adh:.0f}%</b></span>
                </div>
                """, unsafe_allow_html=True)

                d_tbl = d_data[["Store No", "Store Name", "SOS", "SOS Status", "Adoption %", "Make Ahead %", "Waste %", "Pre-Bump %", "Adherence %"]].copy()
                d_tbl = d_tbl.sort_values("Store No", ascending=True)

                d_raw_sos = d_tbl["SOS"].copy()
                d_raw_adopt = d_tbl["Adoption %"].copy()
                d_raw_ma = d_tbl["Make Ahead %"].copy()
                d_raw_waste = d_tbl["Waste %"].copy()
                d_raw_pb = d_tbl["Pre-Bump %"].copy()
                d_raw_adh = d_tbl["Adherence %"].copy()

                d_tbl["SOS"] = d_tbl["SOS"].apply(fmt_sos)
                d_tbl["Adoption %"] = d_tbl["Adoption %"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
                d_tbl["Make Ahead %"] = d_tbl["Make Ahead %"].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
                d_tbl["Waste %"] = d_tbl["Waste %"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "-")
                d_tbl["Pre-Bump %"] = d_tbl["Pre-Bump %"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "-")
                d_tbl["Adherence %"] = d_raw_adh.apply(lambda x: f"{x:.0f}%" if pd.notna(x) else "-")
                d_tbl = d_tbl.rename(columns={"SOS Status": "Status", "Store No": "Store #"}).reset_index(drop=True)
                d_raw_sos = d_raw_sos.reset_index(drop=True)
                d_raw_adopt = d_raw_adopt.reset_index(drop=True)
                d_raw_ma = d_raw_ma.reset_index(drop=True)
                d_raw_waste = d_raw_waste.reset_index(drop=True)
                d_raw_pb = d_raw_pb.reset_index(drop=True)
                d_raw_adh = d_raw_adh.reset_index(drop=True)

                def style_district_row(row):
                    idx = row.name
                    cols = list(row.index)
                    styles = [""] * len(row)
                    if pd.notna(d_raw_sos.get(idx)) and d_raw_sos[idx] > 10:
                        styles[cols.index("SOS")] += f"; {OFF_GUIDE}"
                    if pd.notna(d_raw_adopt.get(idx)) and d_raw_adopt[idx] < 85:
                        styles[cols.index("Adoption %")] += f"; {OFF_GUIDE}"
                    if pd.notna(d_raw_ma.get(idx)) and d_raw_ma[idx] > 0:
                        styles[cols.index("Make Ahead %")] += f"; {OFF_GUIDE}"
                    if pd.notna(d_raw_waste.get(idx)) and d_raw_waste[idx] > 0:
                        styles[cols.index("Waste %")] += f"; {OFF_GUIDE}"
                    if pd.notna(d_raw_pb.get(idx)) and d_raw_pb[idx] > 0.5:
                        styles[cols.index("Pre-Bump %")] += f"; {OFF_GUIDE}"
                    if pd.notna(d_raw_adh.get(idx)) and d_raw_adh[idx] < 85:
                        styles[cols.index("Adherence %")] += f"; {OFF_GUIDE}"
                    return styles

                st.dataframe(d_tbl.style.apply(style_district_row, axis=1), use_container_width=True, hide_index=True)

        # ════════════════════
        # TAB: REPEAT OFFENDERS
        # ════════════════════
        with tab_offenders:
            st.markdown('<div style="font-weight:700; color:#1A3C34; font-size:1.1rem; margin-bottom:0.3rem;">Repeat Offenders</div>', unsafe_allow_html=True)
            st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Stores consistently slow (SOS &gt; 10 min) or non-adherent (&lt;60% overall) across multiple weeks.</p>', unsafe_allow_html=True)

            # Classify each store in each period
            all_classified = kds_raw.copy()
            all_classified["District"] = all_classified["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("Unassigned")
            all_classified["Is Slow"] = all_classified["SOS"] > 10
            all_classified["Is Non-Adherent"] = all_classified["Adherence %"] < 60

            # Slow repeat offenders
            slow_counts = all_classified[all_classified["Is Slow"]].groupby("Store No").agg(
                Times_Slow=("Period", "count"),
                Weeks=("Period", lambda x: ", ".join(sorted(x.unique(), key=lambda p: (int(p[1]), int(p[3]))))),
                Avg_SOS=("SOS", "mean"),
                Store_Name=("Store Name", "first"),
            ).reset_index().sort_values("Times_Slow", ascending=False)
            slow_counts["District"] = slow_counts["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("")

            # Non-adherent repeat offenders
            nonadh_counts = all_classified[all_classified["Is Non-Adherent"]].groupby("Store No").agg(
                Times_Below=("Period", "count"),
                Weeks=("Period", lambda x: ", ".join(sorted(x.unique(), key=lambda p: (int(p[1]), int(p[3]))))),
                Avg_Adherence=("Adherence %", "mean"),
                Store_Name=("Store Name", "first"),
            ).reset_index().sort_values("Times_Below", ascending=False)
            nonadh_counts["District"] = nonadh_counts["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("")

            col_slow, col_nonadh = st.columns(2)
            with col_slow:
                st.markdown(f"""<div style="background:#DC2626; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0;">
                    <span style="font-weight:700;">Slow SOS Repeat Offenders</span>
                    <span style="float:right; font-size:0.82rem;">SOS &gt; 10 min</span>
                </div>""", unsafe_allow_html=True)

                if len(slow_counts) > 0:
                    slow_display = slow_counts[["Store No", "Store_Name", "District", "Times_Slow", "Avg_SOS", "Weeks"]].copy()
                    slow_display["Avg_SOS"] = slow_display["Avg_SOS"].apply(fmt_sos)
                    slow_display = slow_display.rename(columns={"Store_Name": "Store", "Times_Slow": "Times Slow", "Avg_SOS": "Avg SOS", "Weeks": "Weeks Flagged"})
                    st.dataframe(slow_display, use_container_width=True, hide_index=True)
                else:
                    st.success("No repeat slow offenders!")

            with col_nonadh:
                st.markdown(f"""<div style="background:#D97706; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0;">
                    <span style="font-weight:700;">Low Adherence Repeat Offenders</span>
                    <span style="float:right; font-size:0.82rem;">Adherence &lt; 60%</span>
                </div>""", unsafe_allow_html=True)

                if len(nonadh_counts) > 0:
                    nonadh_display = nonadh_counts[["Store No", "Store_Name", "District", "Times_Below", "Avg_Adherence", "Weeks"]].copy()
                    nonadh_display["Avg_Adherence"] = nonadh_display["Avg_Adherence"].apply(lambda x: f"{x:.0f}%")
                    nonadh_display = nonadh_display.rename(columns={"Store_Name": "Store", "Times_Below": "Times Flagged", "Avg_Adherence": "Avg Adherence", "Weeks": "Weeks Flagged"})
                    st.dataframe(nonadh_display, use_container_width=True, hide_index=True)
                else:
                    st.success("No repeat low-adherence offenders!")

            # Repeat offender frequency chart
            st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
            if len(slow_counts) > 0:
                top_slow = slow_counts.head(15).copy()
                top_slow["Label"] = top_slow["Store No"].astype(str) + " - " + top_slow["Store_Name"]
                fig_repeat = go.Figure(go.Bar(
                    x=top_slow["Label"], y=top_slow["Times_Slow"], text=top_slow["Times_Slow"].astype(int),
                    textposition="outside", marker_color="#DC2626",
                ))
                repeat_layout = {**CHART_LAYOUT,
                                 "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Weeks Slow", dtick=1),
                                 "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category", tickfont=dict(size=9))}
                fig_repeat.update_layout(**repeat_layout, title="Slow SOS Frequency (All Weeks)", height=380, xaxis_tickangle=-45)
                st.plotly_chart(fig_repeat, use_container_width=True, config=CHART_CONFIG)

        # ════════════════════
        # TAB: HISTORICAL
        # ════════════════════
        with tab_historical:
            # Quarter filter
            qcol1, qcol2 = st.columns([3, 1])
            with qcol2:
                quarter_opts = ["All Quarters"]
                # Detect which quarters exist in the data
                all_period_nums = sorted(set(int(p[1]) for p in periods_sorted))
                q_map = {"Q1 (P1–P3)": [1, 2, 3], "Q2 (P4–P6)": [4, 5, 6], "Q3 (P7–P9)": [7, 8, 9], "Q4 (P10–P13)": [10, 11, 12, 13]}
                for qlabel, qperiods in q_map.items():
                    if any(pn in qperiods for pn in all_period_nums):
                        quarter_opts.append(qlabel)
                hist_quarter = st.selectbox("Quarter", quarter_opts, index=0, key="kds_hist_quarter")

            # Filter periods by quarter
            if hist_quarter != "All Quarters":
                q_period_nums = q_map[hist_quarter]
                hist_periods = [p for p in periods_sorted if int(p[1]) in q_period_nums]
            else:
                hist_periods = periods_sorted

            range_label = f"{hist_periods[0]} to {hist_periods[-1]}" if hist_periods else "No data"
            st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.1rem; margin-bottom:0.5rem;">Historical Trend — {range_label}</div>', unsafe_allow_html=True)

            # Overall SOS adherence trend
            hist = []
            for p in hist_periods:
                pw = kds_raw[kds_raw["Period"] == p]
                pw_valid = pw[pw["SOS"].notna()]
                n_pw = len(pw_valid)
                n_adh = pw_valid["SOS"].between(0, 10, inclusive="right").sum()
                n_slow_p = (pw_valid["SOS"] > 10).sum()
                n_fast_p = (pw_valid["SOS"] < 7).sum()
                avg_sos_p = pw_valid["SOS"].mean() if len(pw_valid) > 0 else 0
                avg_adh_p = pw["Adherence %"].mean() if pw["Adherence %"].notna().any() else 0
                hist.append({
                    "Period": p, "Adherent": n_adh, "Slow": n_slow_p, "Fast": n_fast_p,
                    "Adherent %": (n_adh / n_pw * 100) if n_pw > 0 else 0,
                    "Avg SOS": avg_sos_p, "Avg Adherence": avg_adh_p, "Total": n_pw,
                })
            hist_df = pd.DataFrame(hist)

            # Adherence % trend line
            fig_hist_line = go.Figure()
            fig_hist_line.add_trace(go.Scatter(
                x=hist_df["Period"], y=hist_df["Adherent %"],
                mode="lines+markers+text", name="SOS Adherent %",
                text=hist_df["Adherent %"].apply(lambda v: f"{v:.0f}%"), textposition="top center",
                line=dict(color="#059669", width=3), marker=dict(size=10),
            ))
            fig_hist_line.add_trace(go.Scatter(
                x=hist_df["Period"], y=hist_df["Avg Adherence"],
                mode="lines+markers+text", name="Overall Adherence %",
                text=hist_df["Avg Adherence"].apply(lambda v: f"{v:.0f}%"), textposition="bottom center",
                line=dict(color="#0D9488", width=2, dash="dot"), marker=dict(size=8),
            ))
            hist_line_layout = {**CHART_LAYOUT,
                                "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="% Adherent", range=[0, 100]),
                                "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category"),
                                "legend": dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)}
            fig_hist_line.update_layout(**hist_line_layout, title="Adherence Rate Trend", height=380)
            st.plotly_chart(fig_hist_line, use_container_width=True, config=CHART_CONFIG)

            # Stacked bar: Adherent / Slow / Fast per period
            fig_hist_bar = go.Figure()
            for status, color in [("Adherent", "#059669"), ("Slow", "#DC2626"), ("Fast", "#D97706")]:
                fig_hist_bar.add_trace(go.Bar(
                    x=hist_df["Period"], y=hist_df[status], name=status,
                    marker_color=color, text=hist_df[status], textposition="inside",
                ))
            hist_bar_layout = {**CHART_LAYOUT, "barmode": "stack",
                               "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Store Count"),
                               "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category"),
                               "legend": dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)}
            fig_hist_bar.update_layout(**hist_bar_layout, title="SOS Distribution by Week", height=380)
            st.plotly_chart(fig_hist_bar, use_container_width=True, config=CHART_CONFIG)

            # Avg SOS trend
            fig_sos_trend = go.Figure()
            fig_sos_trend.add_trace(go.Scatter(
                x=hist_df["Period"], y=hist_df["Avg SOS"],
                mode="lines+markers+text", name="Avg SOS",
                text=hist_df["Avg SOS"].apply(fmt_sos), textposition="top center",
                line=dict(color="#1A3C34", width=3), marker=dict(size=10),
            ))
            fig_sos_trend.add_hline(y=10, line_dash="dash", line_color="#DC2626", line_width=1.5,
                                    annotation_text="10 min target", annotation_font=dict(color="#DC2626", size=10))
            fig_sos_trend.add_hline(y=7, line_dash="dash", line_color="#D97706", line_width=1,
                                    annotation_text="7 min floor", annotation_font=dict(color="#D97706", size=10))
            sos_trend_layout = {**CHART_LAYOUT,
                                "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Avg SOS (min)"),
                                "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category")}
            fig_sos_trend.update_layout(**sos_trend_layout, title="Average SOS Trend", height=350)
            st.plotly_chart(fig_sos_trend, use_container_width=True, config=CHART_CONFIG)

            # ── Historical detail table ──
            st.markdown('<div style="font-weight:700; color:#1F2937; font-size:1rem; margin:1rem 0 0.5rem 0;">Period Summary</div>', unsafe_allow_html=True)
            hist_display = hist_df[["Period", "Total", "Adherent", "Slow", "Fast", "Adherent %", "Avg SOS", "Avg Adherence"]].copy()
            hist_display["Adherent %"] = hist_display["Adherent %"].apply(lambda x: f"{x:.1f}%")
            hist_display["Avg SOS"] = hist_display["Avg SOS"].apply(fmt_sos)
            hist_display["Avg Adherence"] = hist_display["Avg Adherence"].apply(lambda x: f"{x:.0f}%")
            st.dataframe(hist_display, use_container_width=True, hide_index=True)

    else:
        st.warning("No KDS dinner data found. Place kds_dinner.csv in the data/ folder.")


# ════════════════════════════════
# SCHEDULE GUIDE
# ════════════════════════════════
elif selected_tab == "Schedule Guide":

    sched_file = DATA_DIR / "schedule_guide.csv"
    if sched_file.exists():
        sched_df = pd.read_csv(sched_file)

        periods_avail = sorted(sched_df["Period"].unique().tolist())

        # ── HEADER ──
        st.markdown(f"""
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
            <div>
                <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">Schedule Guide</h2>
                <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                    Weekly sales forecast &amp; staffing hours &nbsp;·&nbsp; {len(DISTRICTS)} districts, {sched_df["Store No"].nunique()} stores
                </p>
            </div>
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1.2rem; border-radius:8px; font-weight:700; font-size:0.9rem; white-space:nowrap;">
                HOURS GUIDE · <span style="color:#B7E4C7;">Hourly staff only (excl. GMs)</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Period selector ──
        pcol1, pcol2 = st.columns([3, 1])
        with pcol2:
            sel_period = st.selectbox("Week", list(reversed(periods_avail)), index=0, key="sched_period", label_visibility="collapsed")

        week_data = sched_df[sched_df["Period"] == sel_period].copy()
        start_dt = week_data["Start Date"].iloc[0] if len(week_data) > 0 else ""
        end_dt = week_data["End Date"].iloc[0] if len(week_data) > 0 else ""

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = int(extract_store_number(selected_store))
            week_data = week_data[week_data["Store No"] == sk_num]
        elif selected_district != "All Districts":
            d_nums = {int(s.split(" - ")[0].strip()) for s in DISTRICTS.get(selected_district, [])}
            week_data = week_data[week_data["Store No"].isin(d_nums)]

        # KPIs
        total_sales = week_data["Sales Forecast"].sum()
        total_hours = week_data["Hours Guide"].sum()
        n_stores = week_data["Store No"].nunique()
        avg_sales = total_sales / n_stores if n_stores else 0
        avg_hours = total_hours / n_stores if n_stores else 0

        # Delta vs previous period
        prev_idx = periods_avail.index(sel_period) - 1 if sel_period in periods_avail and periods_avail.index(sel_period) > 0 else -1
        prev_sales = prev_hours = None
        if prev_idx >= 0:
            prev_p = periods_avail[prev_idx]
            prev_data = sched_df[sched_df["Period"] == prev_p]
            prev_sales = prev_data["Sales Forecast"].sum()
            prev_hours = prev_data["Hours Guide"].sum()

        kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
            <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
            <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
            <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
        </div>"""

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.markdown(kpi_style.format(label="TOTAL SALES FORECAST", value=f"${total_sales:,.0f}", color="#1F2937",
                    sub=f"week of {start_dt}"), unsafe_allow_html=True)
        c2.markdown(kpi_style.format(label="TOTAL HOURS GUIDE", value=f"{total_hours:,.0f}", color="#0D9488",
                    sub="hourly staff only"), unsafe_allow_html=True)
        c3.markdown(kpi_style.format(label="STORES", value=n_stores, color="#1F2937",
                    sub="reporting this week"), unsafe_allow_html=True)
        c4.markdown(kpi_style.format(label="AVG SALES / STORE", value=f"${avg_sales:,.0f}", color="#1F2937",
                    sub="per store average"), unsafe_allow_html=True)
        c5.markdown(kpi_style.format(label="AVG HOURS / STORE", value=f"{avg_hours:,.0f}", color="#0D9488",
                    sub="per store average"), unsafe_allow_html=True)

        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

        # ── Takeaways ──
        st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.05rem; margin:0.5rem 0 0.5rem 0;">{sel_period} Takeaways</div>', unsafe_allow_html=True)
        takeaway_style = '<div style="border-left:4px solid {color}; padding:0.5rem 1rem; margin:0.4rem 0; background:#FAFBFC; border-radius:0 6px 6px 0;">{text}</div>'

        # Overall summary with delta
        delta_text = ""
        if prev_sales is not None:
            sales_delta = total_sales - prev_sales
            sales_pct = (sales_delta / prev_sales * 100) if prev_sales else 0
            arrow = "▲" if sales_delta > 0 else "▼"
            delta_color = "#059669" if sales_delta > 0 else "#DC2626"
            delta_text = f' <span style="color:{delta_color}; font-weight:700;">{arrow} ${abs(sales_delta):,.0f} ({abs(sales_pct):.1f}%) vs {periods_avail[prev_idx]}</span>.'

        st.markdown(takeaway_style.format(color="#1A3C34",
            text=f'<b>${total_sales:,.0f} total forecast</b> across {n_stores} stores, {total_hours:,.0f} hours guided.{delta_text}'),
            unsafe_allow_html=True)

        # Highest & lowest sales stores
        if len(week_data) > 1:
            top_store = week_data.loc[week_data["Sales Forecast"].idxmax()]
            bot_store = week_data.loc[week_data["Sales Forecast"].idxmin()]
            st.markdown(takeaway_style.format(color="#059669",
                text=f'<b>Highest forecast:</b> {int(top_store["Store No"])} {top_store["Store Name"]} — ${int(top_store["Sales Forecast"]):,} sales, {int(top_store["Hours Guide"]):,} hours.'),
                unsafe_allow_html=True)
            st.markdown(takeaway_style.format(color="#D97706",
                text=f'<b>Lowest forecast:</b> {int(bot_store["Store No"])} {bot_store["Store Name"]} — ${int(bot_store["Sales Forecast"]):,} sales, {int(bot_store["Hours Guide"]):,} hours.'),
                unsafe_allow_html=True)

        # ── District-grouped tables ──
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="section-title" style="font-size:1.05rem;">Weekly Detail — {start_dt} to {end_dt}</div>', unsafe_allow_html=True)

        sorted_districts = sorted(week_data["District"].unique())

        def render_district_card(district, container):
            """Render a single district card (header + HTML table) into a Streamlit container."""
            d_data = week_data[week_data["District"] == district].copy()
            d_sales = d_data["Sales Forecast"].sum()
            d_hours = d_data["Hours Guide"].sum()
            d_n = len(d_data)

            rows_html = ""
            for i, (_, r) in enumerate(d_data.sort_values("Store No").iterrows()):
                bg = "#FFFFFF" if i % 2 == 0 else "#F9FAFB"
                rows_html += f"""<tr style="background:{bg};">
                    <td style="padding:0.45rem 0.6rem; border-bottom:1px solid #F1F5F9; color:#374151; font-weight:600; text-align:center;">{int(r['Store No'])}</td>
                    <td style="padding:0.45rem 0.6rem; border-bottom:1px solid #F1F5F9; color:#1F2937; font-size:0.88rem;">{r['Store Name']}</td>
                    <td style="padding:0.45rem 0.6rem; border-bottom:1px solid #F1F5F9; color:#059669; font-weight:700; text-align:right;">${int(r['Sales Forecast']):,}</td>
                    <td style="padding:0.45rem 0.6rem; border-bottom:1px solid #F1F5F9; color:#0D9488; font-weight:700; text-align:right;">{int(r['Hours Guide']):,}</td>
                </tr>"""

            th_style = "padding:0.4rem 0.6rem; font-size:0.72rem; color:#6B7280; font-weight:600; text-transform:uppercase; letter-spacing:0.5px; border-bottom:2px solid #E2E8F0;"
            container.markdown(f"""
            <div style="border:1px solid #E2E8F0; border-radius:8px; overflow:hidden; margin-bottom:1rem; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
                <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 0.8rem;
                            display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:0.3rem;">
                    <span style="font-weight:700; font-size:0.9rem;">{district}</span>
                    <span style="font-size:0.75rem; opacity:0.9;">${d_sales:,.0f} &nbsp;·&nbsp; {d_hours:,.0f} hrs &nbsp;·&nbsp; {d_n} stores</span>
                </div>
                <table style="width:100%; border-collapse:collapse;">
                    <thead>
                        <tr style="background:#F1F5F9;">
                            <th style="{th_style} text-align:center;">Store</th>
                            <th style="{th_style} text-align:left;">Name</th>
                            <th style="{th_style} text-align:right;">Sales</th>
                            <th style="{th_style} text-align:right;">Hours</th>
                        </tr>
                    </thead>
                    <tbody>{rows_html}</tbody>
                </table>
            </div>
            """, unsafe_allow_html=True)

        # Render districts side by side (2 per row)
        for i in range(0, len(sorted_districts), 2):
            cols = st.columns(2, gap="medium")
            render_district_card(sorted_districts[i], cols[0])
            if i + 1 < len(sorted_districts):
                render_district_card(sorted_districts[i + 1], cols[1])

    else:
        st.warning("No schedule data found. Place schedule_guide.csv in the data/ folder.")

# ════════════════════════════════
# INTERNAL QSC EVALUATIONS (dashboard style)
# ════════════════════════════════
elif selected_tab == "Internal QSC Evals":

    eval_file = DATA_DIR / "qsc_evals.csv"
    if eval_file.exists():
        evals = pd.read_csv(eval_file)
        evals["No Eval"] = evals["No Eval"].astype(bool)
        evals["Red Flag"] = evals["Red Flag"].astype(bool)

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = int(extract_store_number(selected_store))
            evals = evals[evals["Store No"] == sk_num]
        elif selected_district != "All Districts":
            d_nums = {int(s.split(" - ")[0].strip()) for s in DISTRICTS.get(selected_district, [])}
            evals = evals[evals["Store No"].isin(d_nums)]

        periods_avail = sorted(evals["Period"].unique().tolist(), key=lambda x: (int(x[1]), int(x[3])))
        latest_eval_period = periods_avail[-1] if periods_avail else ""

        # Maps
        name_map = evals.dropna(subset=["City"]).drop_duplicates("Store No", keep="last").set_index("Store No")["City"]
        district_map = evals.drop_duplicates("Store No", keep="last").set_index("Store No")["District"]

        # ── HEADER ──
        st.markdown(f"""
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
            <div>
                <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">Internal QSC Evaluations</h2>
                <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                    Self-evaluation audit tracking &nbsp;·&nbsp; {len(DISTRICTS)} districts, {evals["Store No"].nunique()} stores
                </p>
            </div>
            <div style="background:#DC2626; color:#FFFFFF; padding:0.5rem 1.2rem; border-radius:8px; font-weight:700; font-size:0.9rem; white-space:nowrap;">
                RED FLAG · <span style="color:#FEF3C7;">Score = 0 or &lt; 1 hr</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Period selector ──
        pcol1, pcol2 = st.columns([3, 1])
        with pcol2:
            sel_eval_period = st.selectbox("Week", list(reversed(periods_avail)), index=0, key="eval_period", label_visibility="collapsed")

        week_evals = evals[evals["Period"] == sel_eval_period].copy()
        start_dt = week_evals["Start Date"].iloc[0] if len(week_evals) > 0 else ""
        end_dt = week_evals["End Date"].iloc[0] if len(week_evals) > 0 else ""

        # Classify each store
        def classify_eval(row):
            if row["No Eval"]:
                return "Missed"
            if row["Red Flag"]:
                return "Red Flag"
            return "Completed"

        week_evals["Eval Status"] = week_evals.apply(classify_eval, axis=1)

        completed = week_evals[~week_evals["No Eval"]]
        n_total = week_evals["Store No"].nunique()
        n_completed = len(completed)
        n_no_eval = int(week_evals["No Eval"].sum())
        n_red = int(week_evals["Red Flag"].sum())
        n_5star = len(completed[completed["Stars"] == 5])
        avg_score = completed["Score"].mean() if len(completed) > 0 else 0
        completion_pct = ((n_total - n_no_eval) / n_total * 100) if n_total > 0 else 0

        # ── KPI Cards ──
        kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
            <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
            <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
            <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
        </div>"""

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.markdown(kpi_style.format(label="STORES EVALUATED", value=f"{n_total - n_no_eval}/{n_total}", color="#1F2937",
                    sub=f"{completion_pct:.0f}% completion"), unsafe_allow_html=True)
        c2.markdown(kpi_style.format(label="5-STAR EVALS", value=n_5star, color="#059669",
                    sub=f"{(n_5star/n_completed*100) if n_completed else 0:.0f}% of completed"), unsafe_allow_html=True)
        red_color = "#DC2626" if n_red > 0 else "#059669"
        c3.markdown(kpi_style.format(label="RED FLAGS", value=n_red, color=red_color,
                    sub="score=0 or <1hr"), unsafe_allow_html=True)
        missed_color = "#D97706" if n_no_eval > 0 else "#059669"
        c4.markdown(kpi_style.format(label="MISSED EVALS", value=n_no_eval, color=missed_color,
                    sub="no evaluation done"), unsafe_allow_html=True)
        c5.markdown(kpi_style.format(label="AVG SCORE", value=f"{avg_score:.0f}", color="#1F2937",
                    sub=f"week of {start_dt}"), unsafe_allow_html=True)

        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

        # ── Charts row: Donut + District Distribution ──
        chart_l, chart_r = st.columns(2)

        with chart_l:
            st.markdown(f"""<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem;">
                <div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">
                    <span style="background:#1A3C34; color:#FFFFFF; padding:2px 10px; border-radius:4px; font-size:0.75rem; font-weight:700;">{sel_eval_period}</span>
                    <span style="font-weight:700; color:#1F2937;">Evaluation Mix</span>
                </div>
            </div>""", unsafe_allow_html=True)

            status_counts = week_evals["Eval Status"].value_counts()
            color_map_eval = {"Completed": "#059669", "Red Flag": "#DC2626", "Missed": "#D97706"}
            labels_e = status_counts.index.tolist()
            values_e = status_counts.values.tolist()
            colors_e = [color_map_eval.get(l, "#CBD5E1") for l in labels_e]

            fig_donut_e = go.Figure(go.Pie(
                labels=labels_e, values=values_e, hole=0.55,
                marker=dict(colors=colors_e),
                textinfo="percent", textfont=dict(size=13, color="#FFFFFF"),
                hovertemplate="%{label}: %{value} stores (%{percent})<extra></extra>",
                sort=False,
            ))
            fig_donut_e.update_layout(
                plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
                font=dict(color=FONT_COLOR, size=11),
                margin=dict(l=10, r=10, t=10, b=10),
                height=320,
                legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05, font=dict(size=11)),
            )
            st.plotly_chart(fig_donut_e, use_container_width=True, config=CHART_CONFIG)

        with chart_r:
            st.markdown(f"""<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem;">
                <div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">
                    <span style="background:#1A3C34; color:#FFFFFF; padding:2px 10px; border-radius:4px; font-size:0.75rem; font-weight:700;">{sel_eval_period}</span>
                    <span style="font-weight:700; color:#1F2937;">Distribution by District</span>
                </div>
            </div>""", unsafe_allow_html=True)

            week_evals["District_Label"] = week_evals["District"].fillna("Unknown")
            dist_eval_status = week_evals.groupby(["District_Label", "Eval Status"]).size().unstack(fill_value=0)
            for col in ["Completed", "Red Flag", "Missed"]:
                if col not in dist_eval_status.columns:
                    dist_eval_status[col] = 0

            fig_dist_e = go.Figure()
            for status, color in [("Completed", "#059669"), ("Red Flag", "#DC2626"), ("Missed", "#D97706")]:
                fig_dist_e.add_trace(go.Bar(
                    x=dist_eval_status.index, y=dist_eval_status[status], name=status,
                    marker_color=color, hovertemplate="%{x}<br>" + status + ": %{y}<extra></extra>",
                ))
            dist_e_layout = {**CHART_LAYOUT, "barmode": "stack",
                             "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category", tickfont=dict(size=9)),
                             "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, dtick=2),
                             "legend": dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5, font=dict(size=10)),
                             "margin": dict(l=40, r=10, t=10, b=60)}
            fig_dist_e.update_layout(**dist_e_layout, height=320)
            st.plotly_chart(fig_dist_e, use_container_width=True, config=CHART_CONFIG)

        # ── Takeaways ──
        st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.05rem; margin:1rem 0 0.5rem 0;">{sel_eval_period} Takeaways</div>', unsafe_allow_html=True)
        takeaway_style = '<div style="border-left:4px solid {color}; padding:0.5rem 1rem; margin:0.4rem 0; background:#FAFBFC; border-radius:0 6px 6px 0;">{text}</div>'

        prev_idx = periods_avail.index(sel_eval_period) - 1 if sel_eval_period in periods_avail and periods_avail.index(sel_eval_period) > 0 else -1
        delta_text = ""
        if prev_idx >= 0:
            prev_p = periods_avail[prev_idx]
            prev_evals = evals[evals["Period"] == prev_p]
            prev_red = int(prev_evals["Red Flag"].sum())
            delta_red = n_red - prev_red
            arrow = "▲" if delta_red > 0 else "▼"
            delta_color = "#DC2626" if delta_red > 0 else "#059669"
            delta_text = f' <span style="color:{delta_color}; font-weight:700;">{arrow} {abs(delta_red)} red flags vs {prev_p}</span>.'

        st.markdown(takeaway_style.format(color="#1A3C34",
            text=f'<b>{completion_pct:.0f}% completion</b> ({n_total - n_no_eval}/{n_total} stores evaluated).{delta_text} {n_5star} five-star, {n_red} red flags, {n_no_eval} missed.'),
            unsafe_allow_html=True)

        week_evals["District_Label"] = week_evals["District"].fillna("Unknown")
        dist_scores = week_evals.groupby("District_Label").apply(
            lambda g: g[~g["No Eval"]]["Score"].mean() if (~g["No Eval"]).any() else 0
        ).sort_values(ascending=False)
        if len(dist_scores) > 0:
            st.markdown(takeaway_style.format(color="#059669",
                text=f'<b>Best District:</b> {dist_scores.index[0]} — avg score {dist_scores.iloc[0]:.0f}.'),
                unsafe_allow_html=True)
        if len(dist_scores) > 1:
            st.markdown(takeaway_style.format(color="#DC2626",
                text=f'<b>Worst District:</b> {dist_scores.index[-1]} — avg score {dist_scores.iloc[-1]:.0f}.'),
                unsafe_allow_html=True)

        missed_stores = week_evals[week_evals["No Eval"]]
        if len(missed_stores) > 0:
            missed_list = ", ".join([f"{int(r['Store No'])} ({r['District_Label']})" for _, r in missed_stores.iterrows()])
            st.markdown(takeaway_style.format(color="#D97706",
                text=f'<b>Missed evaluations:</b> {missed_list}'),
                unsafe_allow_html=True)

        # ── Repeat Offenders ──
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="font-size:1.05rem;">Repeat Offenders</div>', unsafe_allow_html=True)
        st.markdown('<p style="color:#6B7280; font-size:0.82rem;">Stores consistently flagged across multiple weeks. These need immediate attention.</p>', unsafe_allow_html=True)

        red_counts = evals[evals["Red Flag"]].groupby("Store No").agg(
            Times_Flagged=("Period", "count"),
            Weeks=("Period", lambda x: ", ".join(sorted(x.unique(), key=lambda p: (int(p[1]), int(p[3]))))),
            Avg_Score=("Score", "mean"),
        ).reset_index().sort_values("Times_Flagged", ascending=False)
        district_map = evals.drop_duplicates("Store No", keep="last").set_index("Store No")["District"]

        noeval_counts = evals[evals["No Eval"]].groupby("Store No").agg(
            Times_Missed=("Period", "count"),
            Weeks_Missed=("Period", lambda x: ", ".join(sorted(x.unique(), key=lambda p: (int(p[1]), int(p[3]))))),
        ).reset_index().sort_values("Times_Missed", ascending=False)

        col_r, col_n = st.columns(2)
        with col_r:
            st.markdown(f"""
            <div style="background:#DC2626; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0;">
                <span style="font-weight:700;">Red Flag Repeat Offenders</span>
                <span style="float:right; font-size:0.82rem;">Score = 0 or Duration &lt; 1 hour</span>
            </div>
            """, unsafe_allow_html=True)
            if len(red_counts) > 0:
                red_display = red_counts.copy()
                red_display["District"] = red_display["Store No"].map(district_map).fillna("")
                red_display["Avg Score"] = red_display["Avg_Score"].apply(lambda x: f"{x:.0f}" if pd.notna(x) else "")
                red_display = red_display.rename(columns={"Times_Flagged": "Times Flagged", "Weeks": "Weeks Flagged"})
                st.dataframe(
                    red_display[["Store No", "District", "Times Flagged", "Avg Score", "Weeks Flagged"]],
                    use_container_width=True, hide_index=True,
                )
            else:
                st.success("No repeat red flag stores!")

        with col_n:
            st.markdown(f"""
            <div style="background:#D97706; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0;">
                <span style="font-weight:700;">Missing Evaluations</span>
                <span style="float:right; font-size:0.82rem;">No evaluation completed</span>
            </div>
            """, unsafe_allow_html=True)
            if len(noeval_counts) > 0:
                noeval_display = noeval_counts.copy()
                noeval_display["District"] = noeval_display["Store No"].map(district_map).fillna("")
                noeval_display = noeval_display.rename(columns={"Times_Missed": "Times Missed", "Weeks_Missed": "Weeks Missed"})
                st.dataframe(
                    noeval_display[["Store No", "District", "Times Missed", "Weeks Missed"]],
                    use_container_width=True, hide_index=True,
                )
            else:
                st.success("All stores completed evaluations!")

        # ── Weekly Detail: District-grouped tables ──
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="font-size:1.05rem;">Weekly Detail</div>', unsafe_allow_html=True)
        st.markdown(f'<p style="color:#374151; font-size:0.9rem; font-weight:600;">Week: {start_dt} — {end_dt}</p>', unsafe_allow_html=True)

        for district in sorted(week_evals["District_Label"].unique()):
            d_evals = week_evals[week_evals["District_Label"] == district].copy()
            d_completed = d_evals[~d_evals["No Eval"]]
            d_5star = len(d_completed[d_completed["Stars"] == 5])
            d_red = d_evals["Red Flag"].sum()
            d_noeval = d_evals["No Eval"].sum()

            badge = ""
            if d_noeval > 0:
                badge += f' <span style="background:#D97706; padding:2px 8px; border-radius:4px; font-size:0.75rem; margin-left:8px;">{int(d_noeval)} missed</span>'
            if d_red > 0:
                badge += f' <span style="background:#DC2626; padding:2px 8px; border-radius:4px; font-size:0.75rem; margin-left:4px;">{int(d_red)} red flags</span>'

            st.markdown(f"""
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0; margin-top:1rem;
                        display:flex; justify-content:space-between; align-items:center;">
                <span style="font-weight:700; font-size:0.95rem;">{district}{badge}</span>
                <span style="font-size:0.82rem;">Completed: <b>{len(d_completed)}</b> &nbsp;|&nbsp; 5-Star: <b>{d_5star}</b></span>
            </div>
            """, unsafe_allow_html=True)

            display_cols = d_evals[["Store No", "Date", "MOD", "Findings", "Score", "Rating", "Duration"]].copy()
            no_eval_mask = d_evals["No Eval"].fillna(False).astype(bool)
            display_cols.loc[no_eval_mask, "Rating"] = "NO EVALUATION"
            display_cols.loc[no_eval_mask, "MOD"] = "-"
            display_cols.loc[no_eval_mask, "Date"] = "-"
            display_cols.loc[no_eval_mask, "Duration"] = "-"
            display_cols["Findings"] = display_cols["Findings"].apply(lambda x: f"{int(x)}" if pd.notna(x) else "-")
            display_cols["Score"] = d_evals["Score"].apply(lambda x: f"{int(x)}" if pd.notna(x) else "-")
            display_cols["Duration"] = display_cols["Duration"].fillna("-").replace("nan", "-")
            display_cols["Date"] = display_cols["Date"].fillna("-").replace("nan", "-")
            display_cols["MOD"] = display_cols["MOD"].fillna("-").replace("nan", "-")
            display_cols = display_cols.reset_index(drop=True)

            def highlight_evals(row):
                styles = [""] * len(row)
                if row["Rating"] == "NO EVALUATION":
                    styles = ["background-color: #FEF3C7; color: #92400E"] * len(row)
                elif str(row["Score"]).isdigit() and int(row["Score"]) == 0:
                    styles = ["background-color: #FEE2E2; color: #991B1B"] * len(row)
                return styles

            styled = display_cols.style.apply(highlight_evals, axis=1)
            st.dataframe(styled, use_container_width=True, hide_index=True)

        # ── Trend chart ──
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="font-size:1rem;">Trend: Red Flags & Missed Evals per Week</div>', unsafe_allow_html=True)

        trend_data = evals.groupby("Period").agg(
            Red_Flags=("Red Flag", "sum"),
            Missed=("No Eval", "sum"),
            Avg_Score=("Score", lambda x: x.dropna().mean()),
        ).reindex(periods_avail).reset_index()

        fig_trend = go.Figure()
        fig_trend.add_trace(go.Bar(x=trend_data["Period"], y=trend_data["Red_Flags"], name="Red Flags",
                                   marker_color="#DC2626", text=trend_data["Red_Flags"].astype(int), textposition="outside"))
        fig_trend.add_trace(go.Bar(x=trend_data["Period"], y=trend_data["Missed"], name="Missed Evals",
                                   marker_color="#D97706", text=trend_data["Missed"].astype(int), textposition="outside"))
        trend_layout = {**CHART_LAYOUT, "barmode": "group",
                        "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Count", dtick=2),
                        "legend": dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)}
        fig_trend.update_layout(**trend_layout, title="Red Flags & Missed Evals by Week")
        st.plotly_chart(fig_trend, use_container_width=True, config=CHART_CONFIG)

    else:
        st.warning("No QSC evaluation data found. Place qsc_evals.csv in the data/ folder.")

# ════════════════════════════════
# COGS VARIANCE
# ════════════════════════════════
elif selected_tab == "COGS Variance":
    st.markdown('<div class="section-title">COGS Variance Analysis</div>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Food cost tracking by store. COGS Variance = COGS Actual − Food Theoretical. Positive = over theoretical (bad). NBO Variance = NBO Actual − Food Theoretical.</p>', unsafe_allow_html=True)

    cogs_file = DATA_DIR / "cogs_variance.csv"
    if cogs_file.exists():
        cogs_raw = pd.read_csv(cogs_file)
        cogs_periods = sorted(cogs_raw["Period"].unique(), key=lambda x: int(x[1]))

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = extract_store_number(selected_store).lstrip("0")
            cogs_raw = cogs_raw[cogs_raw["Store No"].astype(str) == sk_num]
        elif selected_district != "All Districts":
            d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
            cogs_raw = cogs_raw[cogs_raw["Store No"].astype(str).isin(d_nums)]

        cogs_periods = [p for p in cogs_periods if p in cogs_raw["Period"].unique()]

        # Period selector
        cogs_period_opts = ["All Periods"] + list(reversed(cogs_periods))
        sel_cogs_period = st.selectbox("Select Period", cogs_period_opts, index=1 if len(cogs_period_opts) > 1 else 0, key="cogs_period")
        if sel_cogs_period != "All Periods":
            cogs_view = cogs_raw[cogs_raw["Period"] == sel_cogs_period].copy()
        else:
            cogs_view = cogs_raw.copy()

        # KPIs
        avg_cogs_actual = cogs_view["COGS Actual %"].mean() if cogs_view["COGS Actual %"].notna().any() else 0
        avg_cogs_var = cogs_view["COGS Variance %"].mean() if cogs_view["COGS Variance %"].notna().any() else 0
        avg_nbo = cogs_view["NBO Actual %"].mean() if cogs_view["NBO Actual %"].notna().any() else 0
        avg_nbo_var = cogs_view["NBO Variance %"].mean() if cogs_view["NBO Variance %"].notna().any() else 0
        avg_theo = cogs_view["Food Theoretical %"].mean() if cogs_view["Food Theoretical %"].notna().any() else 0
        n_over_1 = (cogs_view.groupby("Store No")["COGS Variance %"].mean() > 1.0).sum()
        n_negative = (cogs_view.groupby("Store No")["COGS Variance %"].mean() < 0).sum()

        cogs_c = "green" if avg_cogs_var <= 0.5 else ("orange" if avg_cogs_var <= 1.5 else "red")
        nbo_c = "green" if avg_nbo_var <= 0.5 else ("orange" if avg_nbo_var <= 1.5 else "red")

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(kpi_card("Avg COGS Actual", f"{avg_cogs_actual:.2f}%"), unsafe_allow_html=True)
        k2.markdown(kpi_card("Avg COGS Variance", f"{avg_cogs_var:+.2f}%", cogs_c), unsafe_allow_html=True)
        k3.markdown(kpi_card("Avg Food Theoretical", f"{avg_theo:.2f}%"), unsafe_allow_html=True)
        k4.markdown(kpi_card("Avg NBO Variance", f"{avg_nbo_var:+.2f}%", nbo_c), unsafe_allow_html=True)

        k5, k6, k7, k8 = st.columns(4)
        k5.markdown(kpi_card("Avg NBO Actual", f"{avg_nbo:.2f}%"), unsafe_allow_html=True)
        k6.markdown(kpi_card("Stores > 1% Variance", str(n_over_1), "red" if n_over_1 > 5 else ("orange" if n_over_1 > 0 else "green")), unsafe_allow_html=True)
        k7.markdown(kpi_card("Stores Under Theo", str(n_negative), "green"), unsafe_allow_html=True)
        k8.markdown(kpi_card("Stores Tracked", str(cogs_view["Store No"].nunique())), unsafe_allow_html=True)

        st.markdown("")

        # ── COGS Variance by Store (bar chart) ──
        st.markdown('<div class="section-title">COGS Variance % by Store</div>', unsafe_allow_html=True)
        store_cogs = cogs_view.groupby(["Store No", "Store Name"]).agg(
            cogs_var=("COGS Variance %", "mean"),
        ).reset_index().sort_values("cogs_var", ascending=False)
        store_cogs["Label"] = store_cogs["Store No"].astype(str) + " - " + store_cogs["Store Name"]

        cogs_colors = [RED if v > 1.5 else (ORANGE if v > 0.5 else GREEN) for v in store_cogs["cogs_var"]]
        fig_cogs = go.Figure(go.Bar(
            x=store_cogs["Label"], y=store_cogs["cogs_var"],
            marker_color=cogs_colors,
            text=store_cogs["cogs_var"].apply(lambda v: f"{v:+.2f}%"),
            textposition="outside",
            hovertemplate="%{x}<br>COGS Variance: %{y:+.2f}%<extra></extra>",
        ))
        fig_cogs.add_hline(y=0, line_color="#374151", line_width=1)
        fig_cogs.add_hline(y=1.0, line_dash="dash", line_color=RED, line_width=1.5,
                           annotation_text="1% threshold", annotation_font=dict(color="#DC2626", size=10))
        cogs_layout = {**CHART_LAYOUT,
                       "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="COGS Variance %"),
                       "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category", tickfont=dict(size=9))}
        fig_cogs.update_layout(**cogs_layout, height=420, xaxis_tickangle=-45)
        st.plotly_chart(fig_cogs, use_container_width=True, config=CHART_CONFIG)

        # ── NBO vs COGS Actual scatter ──
        st.markdown('<div class="section-title">NBO Actual vs COGS Actual</div>', unsafe_allow_html=True)
        scatter_data = cogs_view.groupby(["Store No", "Store Name"]).agg(
            nbo=("NBO Actual %", "mean"),
            cogs=("COGS Actual %", "mean"),
            nbo_var=("NBO Variance %", "mean"),
        ).reset_index()
        scatter_data["Label"] = scatter_data["Store No"].astype(str) + " - " + scatter_data["Store Name"]

        fig_scatter = px.scatter(scatter_data, x="nbo", y="cogs", hover_name="Label",
                                 color_discrete_sequence=[TEAL], size_max=10)
        fig_scatter.update_traces(marker=dict(size=10, opacity=0.7))
        # Add diagonal reference line (NBO = COGS means no difference from adjustments)
        min_val = min(scatter_data["nbo"].min(), scatter_data["cogs"].min()) - 1
        max_val = max(scatter_data["nbo"].max(), scatter_data["cogs"].max()) + 1
        fig_scatter.add_trace(go.Scatter(x=[min_val, max_val], y=[min_val, max_val],
                                         mode="lines", line=dict(dash="dash", color="#9CA3AF"),
                                         showlegend=False, hoverinfo="skip"))
        scatter_layout = {**CHART_LAYOUT,
                          "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="NBO Actual %"),
                          "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="COGS Actual %")}
        fig_scatter.update_layout(**scatter_layout, height=400, title="NBO vs COGS (points below line = adjustments helping)")
        st.plotly_chart(fig_scatter, use_container_width=True, config=CHART_CONFIG)

        # ── Trend over periods ──
        if len(cogs_periods) > 1:
            st.markdown('<div class="section-title">COGS Variance Trend by Period</div>', unsafe_allow_html=True)
            cogs_trend = cogs_raw.groupby("Period").agg(
                cogs_var=("COGS Variance %", "mean"),
                nbo_var=("NBO Variance %", "mean"),
                cogs_actual=("COGS Actual %", "mean"),
            ).reindex(cogs_periods).reset_index()

            fig_ctrend = go.Figure()
            fig_ctrend.add_trace(go.Scatter(x=cogs_trend["Period"], y=cogs_trend["cogs_var"],
                                            mode="lines+markers", name="COGS Variance",
                                            line=dict(color=RED, width=2), marker=dict(size=8)))
            fig_ctrend.add_trace(go.Scatter(x=cogs_trend["Period"], y=cogs_trend["nbo_var"],
                                            mode="lines+markers", name="NBO Variance",
                                            line=dict(color=ORANGE, width=2), marker=dict(size=8)))
            fig_ctrend.add_hline(y=0, line_color="#374151", line_width=1)
            ctrend_layout = {**CHART_LAYOUT,
                             "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Variance %"),
                             "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category"),
                             "legend": dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)}
            fig_ctrend.update_layout(**ctrend_layout, title="Avg COGS & NBO Variance Trend")
            st.plotly_chart(fig_ctrend, use_container_width=True, config=CHART_CONFIG)

        # ── District Summary ──
        st.markdown('<div class="section-title">District COGS Summary</div>', unsafe_allow_html=True)
        cogs_view["District"] = cogs_view["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("Unknown")
        dist_cogs = cogs_view.groupby("District").agg(
            avg_cogs=("COGS Actual %", "mean"),
            avg_var=("COGS Variance %", "mean"),
            avg_nbo=("NBO Actual %", "mean"),
            stores=("Store No", "nunique"),
        ).reset_index().sort_values("avg_var", ascending=False)

        dist_colors = [RED if v > 1.0 else (ORANGE if v > 0.5 else GREEN) for v in dist_cogs["avg_var"]]
        fig_dist = go.Figure(go.Bar(
            x=dist_cogs["District"], y=dist_cogs["avg_var"],
            marker_color=dist_colors,
            text=dist_cogs["avg_var"].apply(lambda v: f"{v:+.2f}%"),
            textposition="outside",
        ))
        fig_dist.add_hline(y=0, line_color="#374151", line_width=1)
        dist_layout = {**CHART_LAYOUT,
                       "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, title="Avg COGS Variance %"),
                       "xaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, type="category")}
        fig_dist.update_layout(**dist_layout, height=350, title="Avg COGS Variance by District")
        st.plotly_chart(fig_dist, use_container_width=True, config=CHART_CONFIG)

        # ── Detail Table ──
        st.markdown('<div class="section-title">Store Detail</div>', unsafe_allow_html=True)
        detail = cogs_view[["Period", "Store No", "Store Name", "NBO Actual %", "Food Theoretical %",
                             "NBO Variance %", "COGS Actual %", "COGS Variance %"]].copy()
        detail["District"] = detail["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("")
        for col in ["NBO Actual %", "Food Theoretical %", "NBO Variance %", "COGS Actual %", "COGS Variance %"]:
            detail[col] = detail[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")
        detail = detail[["Period", "Store No", "Store Name", "District", "NBO Actual %", "Food Theoretical %",
                          "NBO Variance %", "COGS Actual %", "COGS Variance %"]]
        detail = detail.sort_values(["Period", "Store Name"])

        def highlight_cogs(row):
            styles = [""] * len(row)
            try:
                var_val = float(str(row["COGS Variance %"]).replace("%", ""))
                if var_val > 1.5:
                    styles = ["background-color: #FEE2E2; color: #991B1B"] * len(row)
                elif var_val > 0.5:
                    styles = ["background-color: #FEF3C7; color: #92400E"] * len(row)
            except (ValueError, TypeError):
                pass
            return styles

        styled = detail.style.apply(highlight_cogs, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True)

    else:
        st.warning("No COGS data found. Place cogs_variance.csv in the data/ folder.")

# ════════════════════════════════
# SALES PERFORMANCE
# ════════════════════════════════
elif selected_tab == "Sales Performance":
    sales_file = DATA_DIR / "sales_journal.csv"
    if not sales_file.exists():
        st.warning("No sales journal data found. Run parse_sales_journal.py first.")
    else:
        sj = pd.read_csv(sales_file)
        sj["Store No"] = sj["Store No"].astype(str)
        sj["District"] = sj["Store No"].map(STORE_TO_DISTRICT).fillna("Unassigned")

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = extract_store_number(selected_store).lstrip("0")
            sj = sj[sj["Store No"] == sk_num]
        elif selected_district != "All Districts":
            d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
            sj = sj[sj["Store No"].isin(d_nums)]

        # Period selector
        sj_periods = sorted(sj["Period"].unique(), key=lambda x: (int(x[1]), int(x[3])))
        sel_sales_period = st.selectbox("Period", list(reversed(sj_periods)), index=0, key="sales_period")
        sj_week = sj[sj["Period"] == sel_sales_period].copy()

        # Store name lookup
        store_name_map = {}
        for dist, stores in DISTRICTS.items():
            for s in stores:
                snum = s.split(" - ")[0].strip().lstrip("0")
                parts = s.split(" - ")
                store_name_map[snum] = parts[-1].strip() if len(parts) >= 3 else s

        sj_week["Store Name Short"] = sj_week["Store No"].map(store_name_map).fillna("")

        # ── HEADER ──
        date_range = f'{sj_week["Start Date"].iloc[0]} - {sj_week["End Date"].iloc[0]}' if len(sj_week) > 0 else ''
        st.markdown(f"""
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
            <div>
                <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">Sales Performance</h2>
                <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                    FL Wingmen — {len(sj_week)} stores &nbsp;·&nbsp; {sel_sales_period} ({date_range})
                </p>
            </div>
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1.2rem; border-radius:8px; font-weight:700; font-size:0.9rem; white-space:nowrap;">
                {sel_sales_period} · <span style="color:#B7E4C7;">Net Week</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── KPIs ──
        total_net = sj_week["Net Sales"].sum()
        total_gross = sj_week["Gross Sales"].sum()
        total_voids = int(sj_week["Void Count"].sum()) if sj_week["Void Count"].notna().any() else 0
        total_void_sales = sj_week["Void Sales"].sum() if sj_week["Void Sales"].notna().any() else 0
        total_refunds = sj_week["Refund"].sum() if sj_week["Refund"].notna().any() else 0
        total_cos = sj_week["Cash Over/Short"].sum() if sj_week["Cash Over/Short"].notna().any() else 0
        total_checks = int(sj_week["Checks Total"].sum()) if sj_week["Checks Total"].notna().any() else 0
        avg_check = sj_week["Check Avg"].mean() if sj_week["Check Avg"].notna().any() else 0

        kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
            <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
            <div style="color:{color}; font-size:1.8rem; font-weight:800; margin:0.2rem 0;">{value}</div>
            <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
        </div>"""

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(kpi_style.format(label="GROSS SALES", value=f"${total_gross:,.0f}", color="#1F2937",
                    sub=f"incl. tax & comps"), unsafe_allow_html=True)
        c2.markdown(kpi_style.format(label="NET SALES", value=f"${total_net:,.0f}", color="#1F2937",
                    sub=f"{len(sj_week)} stores"), unsafe_allow_html=True)
        c3.markdown(kpi_style.format(label="CHECK AVG", value=f"${avg_check:.2f}", color="#0D9488",
                    sub=f"{total_checks:,} total checks"), unsafe_allow_html=True)
        cos_color = "#DC2626" if (total_cos < -2 or total_cos > 2) else "#059669"
        c4.markdown(kpi_style.format(label="CASH OVER/SHORT", value=f"${total_cos:+,.2f}", color=cos_color,
                    sub="net week"), unsafe_allow_html=True)

        st.markdown("<div style='height:0.3rem;'></div>", unsafe_allow_html=True)

        c5, c6, c7, c8 = st.columns(4)
        c5.markdown(kpi_style.format(label="VOID COUNT", value=f"{total_voids:,}", color="#D97706",
                    sub=f"${total_void_sales:,.2f} void sales"), unsafe_allow_html=True)
        c6.markdown(kpi_style.format(label="REFUNDS", value=f"${abs(total_refunds):,.2f}", color="#DC2626",
                    sub="total refund $"), unsafe_allow_html=True)
        online_total = sj_week["Online Sales"].sum() if sj_week["Online Sales"].notna().any() else 0
        online_pct = (online_total / total_net * 100) if total_net > 0 else 0
        c7.markdown(kpi_style.format(label="ONLINE SALES", value=f"${online_total:,.0f}", color="#0D9488",
                    sub=f"{online_pct:.1f}% of net sales"), unsafe_allow_html=True)
        total_discount = sj_week["Total Discount"].sum() if sj_week["Total Discount"].notna().any() else 0
        disc_pct = (total_discount / total_net * 100) if total_net > 0 else 0
        c8.markdown(kpi_style.format(label="DISCOUNTS", value=f"${total_discount:,.0f}", color="#D97706",
                    sub=f"{disc_pct:.2f}% of net"), unsafe_allow_html=True)

        st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

        # ── Daypart Sales Breakdown ──
        st.markdown('<div class="section-title">Daypart Sales Mix</div>', unsafe_allow_html=True)
        dp_data = []
        for dp in ["Lunch", "Snack", "Dinner", "Late"]:
            dp_sales = sj_week[f"{dp} Sales"].sum() if sj_week[f"{dp} Sales"].notna().any() else 0
            dp_tickets = int(sj_week[f"{dp} Tickets"].sum()) if sj_week[f"{dp} Tickets"].notna().any() else 0
            dp_avg = (dp_sales / dp_tickets) if dp_tickets > 0 else 0
            dp_data.append({"Daypart": dp, "Sales": dp_sales, "Tickets": dp_tickets, "Avg Ticket": dp_avg})

        dp_df = pd.DataFrame(dp_data)
        dp_total = dp_df["Sales"].sum()

        dp_cols = st.columns(4)
        dp_colors = {"Lunch": "#0D9488", "Snack": "#D97706", "Dinner": "#2D6A4F", "Late": "#7C3AED"}
        for i, (_, dpr) in enumerate(dp_df.iterrows()):
            pct = (dpr["Sales"] / dp_total * 100) if dp_total > 0 else 0
            dp_cols[i].markdown(kpi_style.format(
                label=dpr["Daypart"].upper(), value=f"${dpr['Sales']:,.0f}",
                color=dp_colors.get(dpr["Daypart"], "#1F2937"),
                sub=f"{pct:.1f}% · {dpr['Tickets']:,} tickets · ${dpr['Avg Ticket']:.2f} avg"
            ), unsafe_allow_html=True)

        st.markdown("<div style='height:0.3rem;'></div>", unsafe_allow_html=True)

        # Daypart donut chart
        fig_dp = go.Figure(go.Pie(
            labels=dp_df["Daypart"], values=dp_df["Sales"], hole=0.55,
            marker=dict(colors=[dp_colors[dp] for dp in dp_df["Daypart"]]),
            textinfo="percent", textfont=dict(size=13, color="#FFFFFF"),
            hovertemplate="%{label}<br>$%{value:,.0f} (%{percent})<extra></extra>",
        ))
        fig_dp.update_layout(plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG, font=dict(color=FONT_COLOR, size=11),
                             margin=dict(l=10, r=10, t=10, b=10), height=280,
                             legend=dict(orientation="h", yanchor="bottom", y=-0.15, xanchor="center", x=0.5))
        st.plotly_chart(fig_dp, use_container_width=True, config=CHART_CONFIG, key="sales_dp_donut")

        st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

        # ── Charts ──
        st.markdown('<div class="section-title">Net Sales by Store</div>', unsafe_allow_html=True)
        s_sorted = sj_week.sort_values("Net Sales", ascending=False)
        fig_ns = go.Figure(go.Bar(
            x=s_sorted["Store Name Short"], y=s_sorted["Net Sales"],
            marker_color=TEAL,
            hovertemplate="Store %{x}<br>Net Sales: $%{y:,.0f}<extra></extra>",
        ))
        fig_ns.update_layout(**CHART_LAYOUT, height=380, yaxis_title="Net Sales ($)", xaxis_tickangle=-45)
        st.plotly_chart(fig_ns, use_container_width=True, key="sales_ns", config=CHART_CONFIG)

        # ── District-grouped tables ──
        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="font-size:1.05rem;">Sales Detail by District</div>', unsafe_allow_html=True)

        for district in sorted(sj_week["District"].unique()):
            d_data = sj_week[sj_week["District"] == district].copy()
            d_net = d_data["Net Sales"].sum()
            d_checks = int(d_data["Checks Total"].sum()) if d_data["Checks Total"].notna().any() else 0
            d_voids = int(d_data["Void Count"].sum()) if d_data["Void Count"].notna().any() else 0
            d_cos = d_data["Cash Over/Short"].sum() if d_data["Cash Over/Short"].notna().any() else 0

            st.markdown(f"""
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0; margin-top:1rem;
                        display:flex; justify-content:space-between; align-items:center;">
                <span style="font-weight:700; font-size:0.95rem;">{district}</span>
                <span style="font-size:0.82rem;">Net: <b>${d_net:,.0f}</b> &nbsp;|&nbsp; Checks: <b>{d_checks:,}</b> &nbsp;|&nbsp; Voids: <b>{d_voids}</b> &nbsp;|&nbsp; Cash O/S: <b>${d_cos:+,.2f}</b></span>
            </div>
            """, unsafe_allow_html=True)

            dtbl = d_data[["Store No", "Store Name Short", "Gross Sales", "Net Sales", "Checks Total", "Check Avg",
                           "Void Count", "Void Sales", "Refund", "Cash Over/Short", "Online Sales"]].copy()
            dtbl = dtbl.sort_values("Store No", key=lambda x: x.astype(int))

            # Online % column
            dtbl["Online %"] = dtbl.apply(lambda r: (r["Online Sales"] / r["Net Sales"] * 100) if pd.notna(r["Online Sales"]) and pd.notna(r["Net Sales"]) and r["Net Sales"] > 0 else None, axis=1)

            # Keep raw cash O/S for styling
            raw_cos = dtbl["Cash Over/Short"].reset_index(drop=True)

            dtbl.columns = ["Store #", "Store", "Gross Sales", "Net Sales", "Checks", "Check Avg",
                            "Voids", "Void $", "Refund $", "Cash O/S", "Online $", "Online %"]
            dtbl["Net Sales"] = dtbl["Net Sales"].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "-")
            dtbl["Gross Sales"] = dtbl["Gross Sales"].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "-")
            dtbl["Check Avg"] = dtbl["Check Avg"].apply(lambda x: f"${x:.2f}" if pd.notna(x) else "-")
            dtbl["Void $"] = dtbl["Void $"].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "-")
            dtbl["Refund $"] = dtbl["Refund $"].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "-")
            dtbl["Cash O/S"] = dtbl["Cash O/S"].apply(lambda x: f"${x:+,.2f}" if pd.notna(x) else "-")
            dtbl["Online $"] = dtbl["Online $"].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "-")
            dtbl["Online %"] = dtbl["Online %"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
            dtbl["Checks"] = dtbl["Checks"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "-")
            dtbl["Voids"] = dtbl["Voids"].apply(lambda x: f"{int(x)}" if pd.notna(x) else "-")
            dtbl = dtbl.reset_index(drop=True)

            def style_sales_row(row):
                idx = row.name
                cols = list(row.index)
                styles = [""] * len(row)
                cos_val = raw_cos.get(idx)
                if pd.notna(cos_val) and (cos_val < -2 or cos_val > 2):
                    styles[cols.index("Cash O/S")] = "color: #DC2626; font-weight: 700"
                return styles

            st.dataframe(dtbl.style.apply(style_sales_row, axis=1), use_container_width=True, hide_index=True)

# ════════════════════════════════
# LABOR DASHBOARD
# ════════════════════════════════
elif selected_tab == "Labor Dashboard":
    forecast_path = DATA_DIR / "forecast.xlsm"
    if not forecast_path.exists():
        st.warning("No forecast data found. Place forecast.xlsm in the data/ folder.")
    else:
        @st.cache_data(ttl=300)
        def load_labor_data():
            raw = pd.read_excel(DATA_DIR / "forecast.xlsm", sheet_name="Forecast_Data")
            return raw[raw["year"] == 2026].copy()

        labor_raw = load_labor_data()
        if labor_raw.empty:
            st.warning("No 2026 labor data found in forecast file.")
        else:
            available_periods = sorted(labor_raw["period"].unique())
            period_labels = {p: f"Period {p}" for p in available_periods}

            fq1, fq2, fq3 = st.columns(3)
            with fq1:
                quarter_options = ["All Quarters", "Q1 (P1-P3)", "Q2 (P4-P6)"]
                selected_quarter = st.selectbox("Quarter", quarter_options, key="labor_quarter")
            q_periods = {"Q1 (P1-P3)": [1, 2, 3], "Q2 (P4-P6)": [4, 5, 6]}
            if selected_quarter in q_periods:
                qtr_filtered = labor_raw[labor_raw["period"].isin(q_periods[selected_quarter])]
                qtr_available = sorted(p for p in available_periods if p in q_periods[selected_quarter])
            else:
                qtr_filtered = labor_raw
                qtr_available = available_periods

            with fq2:
                period_options = ["All Periods"] + qtr_available
                selected_period = st.selectbox(
                    "Period", period_options,
                    format_func=lambda p: "All Periods" if p == "All Periods" else period_labels.get(p, str(p)),
                    index=len(period_options) - 1,
                    key="labor_period",
                )
            with fq3:
                if selected_period == "All Periods":
                    week_choices = sorted(qtr_filtered["week_d"].unique())
                else:
                    week_choices = sorted(qtr_filtered[qtr_filtered["period"] == selected_period]["week_d"].unique())
                week_options = ["All Weeks"] + list(week_choices)
                selected_labor_week = st.selectbox("Week", week_options, key="labor_week")

            if selected_quarter in q_periods:
                lf = labor_raw[labor_raw["period"].isin(q_periods[selected_quarter])].copy()
            else:
                lf = labor_raw.copy()
            if selected_period != "All Periods":
                lf = lf[lf["period"] == selected_period]
            if selected_labor_week != "All Weeks":
                lf = lf[lf["week_d"] == selected_labor_week]

            lf["store_num"] = lf["store"].apply(forecast_store_num)
            lf["short_name"] = lf["store"].apply(forecast_short_name)
            lf["config_district"] = lf["store_num"].map(STORE_TO_DISTRICT)

            if selected_store != "All Stores":
                sk_num = extract_store_number(selected_store)
                lf = lf[lf["store_num"] == sk_num]
            elif selected_district != "All Districts":
                d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
                lf = lf[lf["store_num"].isin(d_nums)]

            labor_df = lf.groupby(["store_num", "short_name", "config_district"]).agg(
                forecast_sales=("forecast_sales", "sum"),
                actual_sales=("actual_sales", "sum"),
                guide_hours=("guide_hours", "sum"),
                scheduled_hours=("schedule_hours", "sum"),
                actual_hours=("actual_crew_hours", "sum"),
                overtime_hours=("ovt_hours", "sum"),
                schedule_labor=("schedule_labor", "sum"),
                actual_labor_cost=("actual_labor", "sum"),
            ).reset_index()

            labor_df["actual_labor_pct"] = labor_df["actual_labor_cost"] / labor_df["actual_sales"]
            labor_df["schedule_labor_pct"] = labor_df["schedule_labor"] / labor_df["forecast_sales"]
            labor_df["labor_variance"] = labor_df["actual_labor_pct"] - labor_df["schedule_labor_pct"]
            labor_df["hours_variance"] = labor_df["actual_hours"] - labor_df["guide_hours"]

            period_display = f"Period {selected_period}" if selected_period != "All Periods" else "All Periods"
            week_display = selected_labor_week if selected_labor_week != "All Weeks" else ""
            time_display = f"{period_display} {week_display}".strip()

            # ── HEADER ──
            st.markdown(f"""
            <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
                <div>
                    <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">Labor Dashboard</h2>
                    <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                        FL Wingmen — {time_display} &nbsp;·&nbsp; {len(labor_df)} stores reporting
                    </p>
                </div>
                <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1.2rem; border-radius:8px; font-weight:700; font-size:0.9rem; white-space:nowrap;">
                    TARGET · <span style="color:#FFD700;">Labor ≤ 18%</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # ── KPIs ──
            avg_labor = labor_df["actual_labor_pct"].mean()
            avg_var = labor_df["labor_variance"].mean()
            total_ot = labor_df["overtime_hours"].sum()
            total_hours = labor_df["actual_hours"].sum()
            total_cost = labor_df["actual_labor_cost"].sum()

            labor_c = "#059669" if avg_labor <= 0.18 else ("#D97706" if avg_labor <= 0.20 else "#DC2626")
            var_c = "#059669" if avg_var <= 0 else ("#D97706" if avg_var < 0.02 else "#DC2626")
            ot_c = "#059669" if total_ot < 200 else ("#D97706" if total_ot < 400 else "#DC2626")

            kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
                <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
                <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
                <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
            </div>"""

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.markdown(kpi_style.format(label="AVG LABOR %", value=f"{avg_labor:.1%}", color=labor_c,
                        sub="target ≤ 18%"), unsafe_allow_html=True)
            c2.markdown(kpi_style.format(label="LABOR VARIANCE", value=f"{avg_var:+.2%}", color=var_c,
                        sub="actual vs scheduled"), unsafe_allow_html=True)
            c3.markdown(kpi_style.format(label="TOTAL OT HOURS", value=f"{total_ot:,.0f}", color=ot_c,
                        sub="overtime exposure"), unsafe_allow_html=True)
            c4.markdown(kpi_style.format(label="TOTAL CREW HOURS", value=f"{total_hours:,.0f}", color="#1F2937",
                        sub="all stores"), unsafe_allow_html=True)
            c5.markdown(kpi_style.format(label="TOTAL LABOR COST", value=f"${total_cost:,.0f}", color="#1F2937",
                        sub="actual spend"), unsafe_allow_html=True)

            st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

            # ── Takeaways ──
            st.markdown(f'<div style="font-weight:700; color:#1A3C34; font-size:1.05rem; margin:0.5rem 0 0.5rem 0;">{time_display} Takeaways</div>', unsafe_allow_html=True)
            takeaway_style = '<div style="border-left:4px solid {color}; padding:0.5rem 1rem; margin:0.4rem 0; background:#FAFBFC; border-radius:0 6px 6px 0;">{text}</div>'

            n_over = (labor_df["actual_labor_pct"] > 0.18).sum()
            n_under = (labor_df["actual_labor_pct"] <= 0.18).sum()
            n_critical = (labor_df["actual_labor_pct"] > 0.20).sum()
            summary_color = "#059669" if n_over == 0 else ("#D97706" if n_over < len(labor_df) / 2 else "#DC2626")
            st.markdown(takeaway_style.format(color="#1A3C34",
                text=f'<b>{avg_labor:.1%} avg labor</b> across {len(labor_df)} stores — <b>{n_under}</b> at or below 18% target, <b>{n_over}</b> above' + (f' (<span style="color:#DC2626; font-weight:700;">{n_critical} critical &gt;20%</span>).' if n_critical > 0 else '.')),
                unsafe_allow_html=True)

            var_arrow = "▲" if avg_var > 0 else "▼"
            var_text_color = "#DC2626" if avg_var > 0 else "#059669"
            st.markdown(takeaway_style.format(color=var_text_color,
                text=f'<b>Variance:</b> <span style="color:{var_text_color}; font-weight:700;">{var_arrow} {avg_var:+.2%}</span> avg actual vs scheduled. Total OT: <b>{total_ot:,.0f} hours</b> (${total_cost:,.0f} total labor cost).'),
                unsafe_allow_html=True)

            if len(labor_df) > 1:
                worst = labor_df.loc[labor_df["actual_labor_pct"].idxmax()]
                best = labor_df.loc[labor_df["actual_labor_pct"].idxmin()]
                st.markdown(takeaway_style.format(color="#DC2626",
                    text=f'<b>Highest labor:</b> Store {worst["store_num"]} {worst["short_name"]} — {worst["actual_labor_pct"]:.1%} labor, {worst["overtime_hours"]:.0f} OT hrs.'),
                    unsafe_allow_html=True)
                st.markdown(takeaway_style.format(color="#059669",
                    text=f'<b>Lowest labor:</b> Store {best["store_num"]} {best["short_name"]} — {best["actual_labor_pct"]:.1%} labor, {best["overtime_hours"]:.0f} OT hrs.'),
                    unsafe_allow_html=True)

            st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

            # ── Charts ──
            st.markdown('<div class="section-title">Labor % by Store (vs Schedule)</div>', unsafe_allow_html=True)
            lb_sorted = labor_df.sort_values("actual_labor_pct", ascending=False)
            lb_colors = [RED if v > 0.20 else (ORANGE if v > 0.18 else GREEN) for v in lb_sorted["actual_labor_pct"]]
            fig_lb = go.Figure(go.Bar(
                x=lb_sorted["short_name"], y=lb_sorted["actual_labor_pct"] * 100,
                marker_color=lb_colors,
                hovertemplate="%{x}<br>Labor: %{y:.1f}%<extra></extra>",
            ))
            fig_lb.add_hline(y=18, line_dash="dash", line_color=RED, line_width=1.5,
                             annotation_text="18% target", annotation_font=dict(color="#DC2626", size=10))
            fig_lb.update_layout(**CHART_LAYOUT, height=380, yaxis_title="Labor %", xaxis_tickangle=-45)
            st.plotly_chart(fig_lb, use_container_width=True, key="labor_pct", config=CHART_CONFIG)

            ll, lr = st.columns(2)
            with ll:
                st.markdown('<div class="section-title">Labor Variance % by Store</div>', unsafe_allow_html=True)
                lv_sorted = labor_df.sort_values("labor_variance", ascending=False)
                lv_colors = [RED if v > 0.02 else (ORANGE if v > 0 else GREEN) for v in lv_sorted["labor_variance"]]
                fig_lv = go.Figure(go.Bar(
                    x=lv_sorted["short_name"], y=lv_sorted["labor_variance"] * 100,
                    marker_color=lv_colors,
                    hovertemplate="%{x}<br>Variance: %{y:+.2f}%<extra></extra>",
                ))
                fig_lv.add_hline(y=0, line_color="#BDBDBD", line_width=1)
                fig_lv.update_layout(**CHART_LAYOUT, height=370, yaxis_title="Labor Variance %", xaxis_tickangle=-45)
                st.plotly_chart(fig_lv, use_container_width=True, key="labor_var_chart", config=CHART_CONFIG)

            with lr:
                st.markdown('<div class="section-title">Overtime Hours by Store</div>', unsafe_allow_html=True)
                ot_sorted = labor_df.sort_values("overtime_hours", ascending=True)
                ot_colors = [RED if v > 25 else (ORANGE if v > 10 else GREEN) for v in ot_sorted["overtime_hours"]]
                fig_ot = go.Figure(go.Bar(
                    y=ot_sorted["short_name"], x=ot_sorted["overtime_hours"],
                    orientation="h", marker_color=ot_colors,
                    text=ot_sorted["overtime_hours"].apply(lambda x: f"{x:.1f}"),
                    textposition="outside", textfont=dict(size=10, color="#374151"),
                    hovertemplate="%{y}<br>OT: %{x:.1f} hrs<extra></extra>",
                ))
                ot_layout = {**CHART_LAYOUT, "margin": dict(l=100, r=40, t=10, b=30),
                             "yaxis": dict(gridcolor=GRID_COLOR, fixedrange=True, tickfont=dict(size=10))}
                fig_ot.update_layout(**ot_layout, height=max(370, len(ot_sorted) * 22), xaxis_title="OT Hours")
                st.plotly_chart(fig_ot, use_container_width=True, key="labor_ot", config=CHART_CONFIG)

            # Weekly labor trend
            if selected_period != "All Periods":
                wk_trend = lf.groupby("week_d").agg(
                    actual_labor=("actual_labor", "sum"),
                    actual_sales=("actual_sales", "sum"),
                    schedule_labor=("schedule_labor", "sum"),
                    forecast_sales=("forecast_sales", "sum"),
                ).reset_index()
                wk_trend["actual_pct"] = wk_trend["actual_labor"] / wk_trend["actual_sales"] * 100
                wk_trend["sched_pct"] = wk_trend["schedule_labor"] / wk_trend["forecast_sales"] * 100

                if len(wk_trend) > 1:
                    st.markdown('<div class="section-title">Weekly Labor % Trend</div>', unsafe_allow_html=True)
                    fig_wt = go.Figure()
                    fig_wt.add_trace(go.Scatter(
                        x=wk_trend["week_d"], y=wk_trend["actual_pct"],
                        name="Actual Labor %", mode="lines+markers",
                        line=dict(color=ORANGE, width=2), marker=dict(size=8),
                    ))
                    fig_wt.add_trace(go.Scatter(
                        x=wk_trend["week_d"], y=wk_trend["sched_pct"],
                        name="Scheduled Labor %", mode="lines+markers",
                        line=dict(color=TEAL, width=2, dash="dash"), marker=dict(size=8),
                    ))
                    fig_wt.update_layout(
                        **CHART_LAYOUT, height=350, yaxis_title="Labor %",
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                                    font=dict(color="#374151")),
                    )
                    st.plotly_chart(fig_wt, use_container_width=True, key="labor_trend", config=CHART_CONFIG)

            # ── District-grouped tables ──
            st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
            st.markdown(f'<div class="section-title" style="font-size:1.05rem;">Labor Detail by District — {time_display}</div>', unsafe_allow_html=True)

            for district in sorted(labor_df["config_district"].dropna().unique()):
                d_data = labor_df[labor_df["config_district"] == district].copy()
                d_labor = d_data["actual_labor_pct"].mean()
                d_ot = d_data["overtime_hours"].sum()
                d_cost = d_data["actual_labor_cost"].sum()
                d_color = "#FFD700" if d_labor <= 0.18 else "#FF6B6B"

                st.markdown(f"""
                <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0; margin-top:1rem;
                            display:flex; justify-content:space-between; align-items:center;">
                    <span style="font-weight:700; font-size:0.95rem;">{district}</span>
                    <span style="font-size:0.82rem;">Labor: <b>{d_labor:.1%}</b> &nbsp;|&nbsp; OT: <b>{d_ot:,.0f} hrs</b> &nbsp;|&nbsp; Cost: <b>${d_cost:,.0f}</b> &nbsp;|&nbsp; Stores: <b>{len(d_data)}</b></span>
                </div>
                """, unsafe_allow_html=True)

                dtbl = d_data[["store_num", "short_name", "actual_sales", "actual_labor_pct", "schedule_labor_pct",
                               "labor_variance", "guide_hours", "actual_hours", "overtime_hours", "actual_labor_cost"]].copy()
                dtbl.columns = ["Store #", "Store", "Actual Sales", "Labor %", "Sched Labor %",
                                "Variance", "Guide Hrs", "Actual Hrs", "OT Hrs", "Labor Cost"]
                dtbl["Actual Sales"] = dtbl["Actual Sales"].apply(lambda x: f"${x:,.0f}")
                dtbl["Labor %"] = dtbl["Labor %"].apply(lambda x: f"{x:.1%}")
                dtbl["Sched Labor %"] = dtbl["Sched Labor %"].apply(lambda x: f"{x:.1%}")
                dtbl["Variance"] = dtbl["Variance"].apply(lambda x: f"{x:+.2%}")
                dtbl["Labor Cost"] = dtbl["Labor Cost"].apply(lambda x: f"${x:,.0f}")
                dtbl["OT Hrs"] = dtbl["OT Hrs"].apply(lambda x: f"{x:.1f}")
                dtbl = dtbl.sort_values("Store")
                st.dataframe(dtbl, use_container_width=True, hide_index=True)

# ════════════════════════════════
# SMG (GUEST SATISFACTION)
# ════════════════════════════════
elif selected_tab == "SMG (Guest Satisfaction)":
    smg_path = DATA_DIR / "smg_surveys.csv"
    if not smg_path.exists():
        st.warning("No SMG data found. Place smg_surveys.csv in the data/ folder.")
    else:
        smg_raw = pd.read_csv(smg_path)
        smg_raw["District"] = smg_raw["Store No"].astype(str).map(STORE_TO_DISTRICT).fillna("Unassigned")
        smg_raw["Satisfaction %"] = 100 - smg_raw["Dissatisfaction %"]
        smg_raw["Order Accuracy %"] = 100 - smg_raw["Inaccurate Order %"]

        # Apply sidebar filters
        if selected_store != "All Stores":
            sk_num = extract_store_number(selected_store).lstrip("0")
            smg_raw = smg_raw[smg_raw["Store No"].astype(str) == sk_num]
        elif selected_district != "All Districts":
            d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
            smg_raw = smg_raw[smg_raw["Store No"].astype(str).isin(d_nums)]

        # ── Header ──
        st.markdown(f"""
        <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
            <div>
                <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.4rem;">Guest Satisfaction (SMG)</h2>
                <p style="color:#6B7280; font-size:0.85rem; margin:0.2rem 0 0 0;">
                    Q2 (3/29/2026 – 6/27/2026) &nbsp;·&nbsp; {len(smg_raw)} stores &nbsp;·&nbsp; {int(smg_raw["Survey Count"].sum()):,} total surveys
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── KPI Cards ──
        avg_sat = smg_raw["Satisfaction %"].mean()
        avg_dissat = smg_raw["Dissatisfaction %"].mean()
        avg_acc = smg_raw["Order Accuracy %"].mean()
        avg_greeted = smg_raw["Greeted with Smile %"].mean()
        total_surveys = int(smg_raw["Survey Count"].sum())

        kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
            <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
            <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
            <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
        </div>"""

        sat_c = "#059669" if avg_sat >= 95 else ("#D97706" if avg_sat >= 90 else "#DC2626")
        dissat_c = "#059669" if avg_dissat <= 5 else ("#D97706" if avg_dissat <= 8 else "#DC2626")
        acc_c = "#059669" if avg_acc >= 97 else ("#D97706" if avg_acc >= 95 else "#DC2626")
        greet_c = "#059669" if avg_greeted >= 95 else ("#D97706" if avg_greeted >= 90 else "#DC2626")

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(kpi_style.format(label="SATISFACTION", value=f"{avg_sat:.1f}%", color=sat_c,
                    sub=f"{total_surveys:,} surveys"), unsafe_allow_html=True)
        k2.markdown(kpi_style.format(label="DISSATISFACTION", value=f"{avg_dissat:.1f}%", color=dissat_c,
                    sub="target ≤ 5%"), unsafe_allow_html=True)
        k3.markdown(kpi_style.format(label="ORDER ACCURACY", value=f"{avg_acc:.1f}%", color=acc_c,
                    sub="target ≥ 97%"), unsafe_allow_html=True)
        k4.markdown(kpi_style.format(label="GREETED WITH SMILE", value=f"{avg_greeted:.1f}%", color=greet_c,
                    sub="target ≥ 95%"), unsafe_allow_html=True)

        st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

        # ── Dissatisfaction Chart ──
        st.markdown('<div class="section-title">Dissatisfaction % by Store (lower is better)</div>', unsafe_allow_html=True)
        dis_sorted = smg_raw.sort_values("Dissatisfaction %", ascending=False)
        dis_colors = [RED if v > 8 else (ORANGE if v > 5 else GREEN) for v in dis_sorted["Dissatisfaction %"]]
        fig_dis = go.Figure(go.Bar(
            x=dis_sorted["Store Name"], y=dis_sorted["Dissatisfaction %"],
            marker_color=dis_colors,
            hovertemplate="%{x}<br>Dissat: %{y:.2f}%<extra></extra>",
        ))
        fig_dis.add_hline(y=5, line_dash="dash", line_color=RED, line_width=1.5,
                          annotation_text="5% target", annotation_font=dict(color="#DC2626", size=10))
        fig_dis.update_layout(**CHART_LAYOUT, height=380, yaxis_title="Dissatisfaction %", xaxis_tickangle=-45)
        st.plotly_chart(fig_dis, use_container_width=True, key="smg_dissat", config=CHART_CONFIG)

        # ── Side-by-side charts ──
        sl, sr = st.columns(2)
        with sl:
            st.markdown('<div class="section-title">Order Accuracy % by Store</div>', unsafe_allow_html=True)
            acc_sorted = smg_raw.sort_values("Order Accuracy %")
            acc_colors = [RED if v < 95 else (ORANGE if v < 97 else GREEN) for v in acc_sorted["Order Accuracy %"]]
            fig_acc = go.Figure(go.Bar(
                x=acc_sorted["Store Name"], y=acc_sorted["Order Accuracy %"],
                marker_color=acc_colors,
                hovertemplate="%{x}<br>Accuracy: %{y:.1f}%<extra></extra>",
            ))
            fig_acc.add_hline(y=97, line_dash="dash", line_color=GREEN, line_width=1.5,
                              annotation_text="97% target", annotation_font=dict(color="#059669", size=10))
            fig_acc.update_layout(**CHART_LAYOUT, height=370, yaxis_title="Accuracy %", xaxis_tickangle=-45)
            st.plotly_chart(fig_acc, use_container_width=True, key="smg_accuracy", config=CHART_CONFIG)

        with sr:
            st.markdown('<div class="section-title">Greeted with Smile % by Store</div>', unsafe_allow_html=True)
            greet_sorted = smg_raw.sort_values("Greeted with Smile %")
            greet_colors = [RED if v < 90 else (ORANGE if v < 95 else GREEN) for v in greet_sorted["Greeted with Smile %"]]
            fig_greet = go.Figure(go.Bar(
                x=greet_sorted["Store Name"], y=greet_sorted["Greeted with Smile %"],
                marker_color=greet_colors,
                hovertemplate="%{x}<br>Greeted: %{y:.1f}%<extra></extra>",
            ))
            fig_greet.add_hline(y=95, line_dash="dash", line_color=GREEN, line_width=1.5,
                                annotation_text="95% target", annotation_font=dict(color="#059669", size=10))
            fig_greet.update_layout(**CHART_LAYOUT, height=370, yaxis_title="Greeted %", xaxis_tickangle=-45)
            st.plotly_chart(fig_greet, use_container_width=True, key="smg_greeted", config=CHART_CONFIG)

        # ── Detail Table by District ──
        st.markdown('<div class="section-title">SMG Detail by District</div>', unsafe_allow_html=True)
        for district in sorted(smg_raw["District"].unique()):
            d_data = smg_raw[smg_raw["District"] == district].copy()
            d_avg_dissat = d_data["Dissatisfaction %"].mean()
            d_avg_acc = d_data["Order Accuracy %"].mean()
            d_surveys = int(d_data["Survey Count"].sum())

            st.markdown(f"""
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0; margin-top:1rem;
                        display:flex; justify-content:space-between; align-items:center;">
                <span style="font-weight:700; font-size:0.95rem;">{district}</span>
                <span style="font-size:0.82rem;">Dissat: <b>{d_avg_dissat:.1f}%</b> &nbsp;|&nbsp; Accuracy: <b>{d_avg_acc:.1f}%</b> &nbsp;|&nbsp; Surveys: <b>{d_surveys:,}</b></span>
            </div>
            """, unsafe_allow_html=True)

            dtbl = d_data[["Store No", "Store Name", "Survey Count", "Dissatisfaction %", "Inaccurate Order %", "Order Accuracy %", "Greeted with Smile %"]].copy()
            dtbl = dtbl.sort_values("Store No")
            dtbl.columns = ["Store #", "Store", "Surveys", "Dissat %", "Inaccurate %", "Accuracy %", "Greeted %"]

            # Style off-guide values
            raw_dissat = dtbl["Dissat %"].copy().reset_index(drop=True)
            raw_inacc = dtbl["Inaccurate %"].copy().reset_index(drop=True)
            raw_greet = dtbl["Greeted %"].copy().reset_index(drop=True)

            dtbl["Dissat %"] = dtbl["Dissat %"].apply(lambda x: f"{x:.2f}%")
            dtbl["Inaccurate %"] = dtbl["Inaccurate %"].apply(lambda x: f"{x:.2f}%")
            dtbl["Accuracy %"] = dtbl["Accuracy %"].apply(lambda x: f"{x:.1f}%")
            dtbl["Greeted %"] = dtbl["Greeted %"].apply(lambda x: f"{x:.1f}%")
            dtbl["Surveys"] = dtbl["Surveys"].apply(lambda x: f"{int(x):,}")
            dtbl = dtbl.reset_index(drop=True)

            OFF_SMG = "color: #DC2626; font-weight: 700"
            def style_smg_row(row):
                idx = row.name
                cols = list(row.index)
                styles = [""] * len(row)
                if pd.notna(raw_dissat.get(idx)) and raw_dissat[idx] > 5:
                    styles[cols.index("Dissat %")] = OFF_SMG
                if pd.notna(raw_inacc.get(idx)) and raw_inacc[idx] > 3:
                    styles[cols.index("Inaccurate %")] = OFF_SMG
                if pd.notna(raw_greet.get(idx)) and raw_greet[idx] < 95:
                    styles[cols.index("Greeted %")] = OFF_SMG
                return styles

            st.dataframe(dtbl.style.apply(style_smg_row, axis=1), use_container_width=True, hide_index=True)

# ════════════════════════════════
# DISTRICT COMPARISON
# ════════════════════════════════
elif selected_tab == "District Comparison":
    import numpy as np

    all_stores = []
    for dist, stores in DISTRICTS.items():
        for s in stores:
            snum = s.split(" - ")[0].strip().lstrip("0")
            sname = s.split("-", 2)[2].strip()[:22] if len(s.split("-", 2)) >= 3 else s[:22]
            all_stores.append({"store": s, "store_num": snum, "short_name": sname, "district": dist})
    comp_df = pd.DataFrame(all_stores)

    # Real labor data from forecast
    forecast_path = DATA_DIR / "forecast.xlsm"
    has_labor = False
    if forecast_path.exists():
        @st.cache_data(ttl=300)
        def load_district_labor():
            raw = pd.read_excel(forecast_path, sheet_name="Forecast_Data")
            return raw[raw["year"] == 2026].copy()
        dc_labor_raw = load_district_labor()
        if not dc_labor_raw.empty:
            has_labor = True
            dc_available_periods = sorted(dc_labor_raw["period"].unique())
            dc_period_labels = {p: f"Period {p}" for p in dc_available_periods}

            dcq1, dcq2, dcq3 = st.columns(3)
            with dcq1:
                dc_quarter_options = ["All Quarters", "Q1 (P1-P3)", "Q2 (P4-P6)"]
                dc_selected_quarter = st.selectbox("Quarter", dc_quarter_options, key="dc_quarter")
            dc_q_periods = {"Q1 (P1-P3)": [1, 2, 3], "Q2 (P4-P6)": [4, 5, 6]}
            if dc_selected_quarter in dc_q_periods:
                dc_qtr_filtered = dc_labor_raw[dc_labor_raw["period"].isin(dc_q_periods[dc_selected_quarter])]
                dc_qtr_available = sorted(p for p in dc_available_periods if p in dc_q_periods[dc_selected_quarter])
            else:
                dc_qtr_filtered = dc_labor_raw
                dc_qtr_available = dc_available_periods

            with dcq2:
                dc_period_options = ["All Periods"] + dc_qtr_available
                dc_selected_period = st.selectbox(
                    "Period", dc_period_options,
                    format_func=lambda p: "All Periods" if p == "All Periods" else dc_period_labels.get(p, str(p)),
                    index=len(dc_period_options) - 1,
                    key="dc_period",
                )
            with dcq3:
                if dc_selected_period == "All Periods":
                    dc_week_choices = sorted(dc_qtr_filtered["week_d"].unique())
                else:
                    dc_week_choices = sorted(dc_qtr_filtered[dc_qtr_filtered["period"] == dc_selected_period]["week_d"].unique())
                dc_week_options = ["All Weeks"] + list(dc_week_choices)
                dc_selected_week = st.selectbox("Week", dc_week_options, key="dc_week")

            dc_labor = dc_labor_raw.copy()
            if dc_selected_quarter in dc_q_periods:
                dc_labor = dc_labor[dc_labor["period"].isin(dc_q_periods[dc_selected_quarter])]
            if dc_selected_period != "All Periods":
                dc_labor = dc_labor[dc_labor["period"] == dc_selected_period]
            if dc_selected_week != "All Weeks":
                dc_labor = dc_labor[dc_labor["week_d"] == dc_selected_week]

            dc_labor["store_num"] = dc_labor["store"].apply(forecast_store_num)
            labor_agg = dc_labor.groupby("store_num").agg(
                actual_sales=("actual_sales", "sum"),
                forecast_sales=("forecast_sales", "sum"),
                actual_labor=("actual_labor", "sum"),
                schedule_labor=("schedule_labor", "sum"),
                ovt_hours=("ovt_hours", "sum"),
            ).reset_index()
            labor_agg["labor_pct"] = labor_agg["actual_labor"] / labor_agg["actual_sales"]
            labor_agg["sched_labor_pct"] = labor_agg["schedule_labor"] / labor_agg["forecast_sales"]
            labor_agg["labor_variance"] = labor_agg["labor_pct"] - labor_agg["sched_labor_pct"]
            labor_map = dict(zip(labor_agg["store_num"].astype(str), labor_agg["labor_pct"]))
            lv_map = dict(zip(labor_agg["store_num"].astype(str), labor_agg["labor_variance"]))
            ot_map = dict(zip(labor_agg["store_num"].astype(str), labor_agg["ovt_hours"]))
            comp_df["labor_pct"] = comp_df["store_num"].map(labor_map)
            comp_df["labor_variance"] = comp_df["store_num"].map(lv_map)
            comp_df["overtime"] = comp_df["store_num"].map(ot_map).fillna(0)

    if not has_labor:
        comp_df["labor_pct"] = None
        comp_df["labor_variance"] = None
        comp_df["overtime"] = 0

    # Real KDS data
    if not daily_df_all.empty:
        latest_date = daily_df_all["data_date"].max()
        daily_latest = daily_df_all[daily_df_all["data_date"] == latest_date]
        sos_map = dict(zip(daily_latest["store_num"].astype(str), daily_latest["sos_min"]))
        comp_df["sos_min"] = comp_df["store_num"].map(sos_map)
    else:
        comp_df["sos_min"] = None

    # Sales & SMG still sample until real data provided
    np.random.seed(42)
    comp_df["net_sales"] = np.random.uniform(28000, 65000, len(comp_df)).round(0)
    comp_df["sss_growth"] = np.random.uniform(-5, 15, len(comp_df)).round(1)
    np.random.seed(77)
    comp_df["osat"] = np.random.uniform(55, 95, len(comp_df)).round(1)

    district_agg = comp_df.groupby("district").agg(
        stores=("store", "count"),
        total_sales=("net_sales", "sum"),
        avg_sales=("net_sales", "mean"),
        avg_sss=("sss_growth", "mean"),
        avg_labor=("labor_pct", "mean"),
        avg_labor_var=("labor_variance", "mean"),
        total_ot=("overtime", "sum"),
        avg_osat=("osat", "mean"),
        avg_sos=("sos_min", "mean"),
    ).reset_index()
    district_agg = district_agg.sort_values("district")

    st.markdown(f'<p style="color:#6B7280; font-size:0.85rem;">District Comparison &nbsp;|&nbsp; {len(district_agg)} districts &nbsp;|&nbsp; <span style="color:#059669; font-weight:600;">Labor &amp; KDS: Real Data</span> &nbsp;|&nbsp; <span style="color:#D97706; font-weight:600;">Sales &amp; SMG: Sample</span></p>', unsafe_allow_html=True)

    dist_colors = [GREEN, TEAL, GOLD, ORANGE, "#7C3AED", "#0EA5E9"]

    st.markdown('<div class="section-title">Total Sales by District</div>', unsafe_allow_html=True)
    fig_ds = go.Figure(go.Bar(
        x=district_agg["district"], y=district_agg["total_sales"],
        marker_color=dist_colors[:len(district_agg)],
        hovertemplate="%{x}<br>Sales: $%{y:,.0f}<extra></extra>",
    ))
    fig_ds.update_layout(**CHART_LAYOUT, height=350, yaxis_title="Total Sales ($)")
    st.plotly_chart(fig_ds, use_container_width=True, key="dist_sales", config=CHART_CONFIG)

    dl, dlv, dm, dr = st.columns(4)
    with dl:
        st.markdown('<div class="section-title">Avg Labor % by District</div>', unsafe_allow_html=True)
        lb_colors = [RED if v > 0.20 else (ORANGE if v > 0.18 else GREEN) for v in district_agg["avg_labor"]]
        fig_dl = go.Figure(go.Bar(
            x=district_agg["district"], y=district_agg["avg_labor"] * 100,
            marker_color=lb_colors,
            hovertemplate="%{x}<br>Labor: %{y:.1f}%<extra></extra>",
        ))
        fig_dl.add_hline(y=18, line_dash="dash", line_color=RED, line_width=1.5)
        fig_dl.update_layout(**CHART_LAYOUT, height=350, yaxis_title="Labor %")
        st.plotly_chart(fig_dl, use_container_width=True, key="dist_labor", config=CHART_CONFIG)

    with dlv:
        st.markdown('<div class="section-title">Labor Variance % by District</div>', unsafe_allow_html=True)
        lv_colors = [RED if v > 0.02 else (ORANGE if v > 0 else GREEN) for v in district_agg["avg_labor_var"]]
        fig_dlv = go.Figure(go.Bar(
            x=district_agg["district"], y=district_agg["avg_labor_var"] * 100,
            marker_color=lv_colors,
            hovertemplate="%{x}<br>Variance: %{y:+.2f}%<extra></extra>",
        ))
        fig_dlv.add_hline(y=0, line_color="#BDBDBD", line_width=1)
        fig_dlv.update_layout(**CHART_LAYOUT, height=350, yaxis_title="Labor Variance %")
        st.plotly_chart(fig_dlv, use_container_width=True, key="dist_labor_var", config=CHART_CONFIG)

    with dm:
        st.markdown('<div class="section-title">Avg OSAT by District</div>', unsafe_allow_html=True)
        osat_colors = [RED if v < 70 else (ORANGE if v < 80 else GREEN) for v in district_agg["avg_osat"]]
        fig_do = go.Figure(go.Bar(
            x=district_agg["district"], y=district_agg["avg_osat"],
            marker_color=osat_colors,
            hovertemplate="%{x}<br>OSAT: %{y:.1f}%<extra></extra>",
        ))
        fig_do.add_hline(y=80, line_dash="dash", line_color=GREEN, line_width=1.5)
        fig_do.update_layout(**CHART_LAYOUT, height=350, yaxis_title="OSAT %")
        st.plotly_chart(fig_do, use_container_width=True, key="dist_osat", config=CHART_CONFIG)

    with dr:
        st.markdown('<div class="section-title">Avg SOS by District</div>', unsafe_allow_html=True)
        sos_colors = [RED if v > 12 else (ORANGE if v > 10 else GREEN) for v in district_agg["avg_sos"]]
        fig_dsos = go.Figure(go.Bar(
            x=district_agg["district"], y=district_agg["avg_sos"],
            marker_color=sos_colors,
            hovertemplate="%{x}<br>SOS: %{text}<extra></extra>",
        ))
        fig_dsos.add_hline(y=10, line_dash="dash", line_color=RED, line_width=1.5)
        fig_dsos.update_layout(**CHART_LAYOUT, height=350, yaxis_title="SOS (min)")
        st.plotly_chart(fig_dsos, use_container_width=True, key="dist_sos", config=CHART_CONFIG)

    st.markdown('<div class="section-title">District Summary Table</div>', unsafe_allow_html=True)
    dtbl = district_agg.copy()
    dtbl.columns = ["District", "Stores", "Total Sales", "Avg Sales/Store", "Avg SSS %", "Avg Labor %", "Labor Variance %", "Total OT Hrs", "Avg OSAT %", "Avg SOS (min)"]
    dtbl["Total Sales"] = dtbl["Total Sales"].apply(lambda x: f"${x:,.0f}")
    dtbl["Avg Sales/Store"] = dtbl["Avg Sales/Store"].apply(lambda x: f"${x:,.0f}")
    dtbl["Avg SSS %"] = dtbl["Avg SSS %"].apply(lambda x: f"{x:+.1f}%")
    dtbl["Avg Labor %"] = dtbl["Avg Labor %"].apply(lambda x: f"{x:.1%}")
    dtbl["Labor Variance %"] = dtbl["Labor Variance %"].apply(lambda x: f"{x:+.2%}" if pd.notna(x) else "—")
    dtbl["Total OT Hrs"] = dtbl["Total OT Hrs"].apply(lambda x: f"{x:.0f}")
    dtbl["Avg OSAT %"] = dtbl["Avg OSAT %"].apply(lambda x: f"{x:.1f}%")
    dtbl["Avg SOS (min)"] = dtbl["Avg SOS (min)"].apply(fmt_sos)
    st.dataframe(dtbl, use_container_width=True, hide_index=True)

# ════════════════════════════════
# SCORECARD
# ════════════════════════════════
elif selected_tab == "Scorecard":
    import numpy as np

    all_stores = []
    for dist, stores in DISTRICTS.items():
        for s in stores:
            snum = s.split(" - ")[0].strip().lstrip("0")
            sname = s.split("-", 2)[2].strip()[:22] if len(s.split("-", 2)) >= 3 else s[:22]
            all_stores.append({"store_num": snum, "short_name": sname, "district": dist})
    sc_df = pd.DataFrame(all_stores)

    # KDS data (latest date)
    if not daily_df_all.empty:
        latest = daily_df_all[daily_df_all["data_date"] == daily_df_all["data_date"].max()]
        kds_map = latest.set_index("store_num")[["sos_min", "pre_bump", "bone_in_adopt", "make_ahead_rate", "waste"]].to_dict("index")
        for col in ["sos_min", "pre_bump", "bone_in_adopt", "make_ahead_rate", "waste"]:
            sc_df[col] = sc_df["store_num"].map(lambda s, c=col: kds_map.get(s, {}).get(c))

    # Labor data
    forecast_path = DATA_DIR / "forecast.xlsm"
    if forecast_path.exists():
        @st.cache_data(ttl=300)
        def load_sc_labor():
            raw = pd.read_excel(forecast_path, sheet_name="Forecast_Data")
            return raw[raw["year"] == 2026].copy()
        sc_labor = load_sc_labor()
        if not sc_labor.empty:
            sc_labor["store_num"] = sc_labor["store"].apply(forecast_store_num)
            la = sc_labor.groupby("store_num").agg(
                actual_sales=("actual_sales", "sum"), forecast_sales=("forecast_sales", "sum"),
                actual_labor=("actual_labor", "sum"), schedule_labor=("schedule_labor", "sum"),
                ovt_hours=("ovt_hours", "sum"),
            ).reset_index()
            la["labor_pct"] = la["actual_labor"] / la["actual_sales"]
            la["sched_labor_pct"] = la["schedule_labor"] / la["forecast_sales"]
            la["labor_var"] = la["labor_pct"] - la["sched_labor_pct"]
            for col in ["labor_pct", "labor_var", "ovt_hours"]:
                sc_df[col] = sc_df["store_num"].map(dict(zip(la["store_num"].astype(str), la[col])))

    # SMG data
    smg_path = DATA_DIR / "smg_q1.xlsx"
    if smg_path.exists():
        @st.cache_data(ttl=300)
        def load_sc_smg():
            raw = pd.read_excel(smg_path, header=None, skiprows=2)
            raw.columns = ["Store", "Measure", "Current", "LastYear", "Difference", "Count", "Count_Current", "Count_LastYear"]
            raw = raw[(raw["Measure"] != "Measure") & (raw["Store"] != "Combined")].copy()
            raw["Current"] = pd.to_numeric(raw["Current"], errors="coerce")
            store_pat = raw["Store"].str.extract(r"^(\d+)\s*-\s*(.*)")
            raw["store_num"] = store_pat[0].str.lstrip("0")
            return raw
        sc_smg = load_sc_smg()
        dissat = sc_smg[sc_smg["Measure"] == "Dissatisfaction"]
        sc_df["dissat_pct"] = sc_df["store_num"].map(dict(zip(dissat["store_num"], dissat["Current"] * 100)))

    # Scoring: each metric gets 0-100 points
    def score_sos(v):
        if pd.isna(v): return None
        if v < 10: return 100
        if v < 13: return 70
        return max(0, 100 - (v - 10) * 10)

    def score_lower_better(v, good, bad):
        if pd.isna(v): return None
        if v <= good: return 100
        if v >= bad: return 0
        return max(0, 100 - (v - good) / (bad - good) * 100)

    def score_higher_better(v, bad, good):
        if pd.isna(v): return None
        if v >= good: return 100
        if v <= bad: return 0
        return (v - bad) / (good - bad) * 100

    sc_df["score_sos"] = sc_df["sos_min"].apply(score_sos) if "sos_min" in sc_df.columns else None
    sc_df["score_prebump"] = sc_df["pre_bump"].apply(lambda v: score_lower_better(v, 0.5, 3)) if "pre_bump" in sc_df.columns else None
    sc_df["score_adopt"] = sc_df["bone_in_adopt"].apply(lambda v: score_higher_better(v, 70, 95)) if "bone_in_adopt" in sc_df.columns else None
    sc_df["score_waste"] = sc_df["waste"].apply(lambda v: score_lower_better(v, 2, 8)) if "waste" in sc_df.columns else None
    sc_df["score_labor"] = sc_df["labor_pct"].apply(lambda v: score_lower_better(v * 100, 18, 24) if pd.notna(v) else None) if "labor_pct" in sc_df.columns else None
    sc_df["score_dissat"] = sc_df["dissat_pct"].apply(lambda v: score_lower_better(v, 3, 12)) if "dissat_pct" in sc_df.columns else None

    score_cols = [c for c in sc_df.columns if c.startswith("score_")]
    sc_df["composite"] = sc_df[score_cols].mean(axis=1).fillna(0)
    sc_df["rank"] = sc_df["composite"].rank(ascending=False, method="min").astype(int)
    sc_df = sc_df.sort_values("composite", ascending=False)

    if selected_store != "All Stores":
        sk_num = extract_store_number(selected_store)
        sc_df = sc_df[sc_df["store_num"] == sk_num]
    elif selected_district != "All Districts":
        d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
        sc_df = sc_df[sc_df["store_num"].isin(d_nums)]

    st.markdown(f'<p style="color:#6B7280; font-size:0.85rem;">Store Scorecard &nbsp;|&nbsp; {len(sc_df)} stores &nbsp;|&nbsp; Composite score across KDS, Labor &amp; SMG</p>', unsafe_allow_html=True)

    # Top 5 / Bottom 5
    if len(sc_df) > 5:
        t5, b5 = st.columns(2)
        with t5:
            st.markdown('<div class="section-title">Top 5 Stores</div>', unsafe_allow_html=True)
            top5 = sc_df.head(5)
            for _, row in top5.iterrows():
                score = row["composite"]
                color = GREEN if score >= 75 else (ORANGE if score >= 50 else RED)
                st.markdown(f'<div style="display:flex; justify-content:space-between; align-items:center; padding:0.5rem 0.8rem; margin:0.3rem 0; background:#FFFFFF; border-left:4px solid {color}; border-radius:4px; border:1px solid #E8ECF0;">'
                            f'<span style="color:#1F2937; font-weight:600;">#{int(row["rank"])} {row["short_name"]}</span>'
                            f'<span style="color:{color}; font-weight:700; font-size:1.1rem;">{score:.0f}</span></div>', unsafe_allow_html=True)
        with b5:
            st.markdown('<div class="section-title">Bottom 5 Stores</div>', unsafe_allow_html=True)
            bot5 = sc_df.tail(5).iloc[::-1]
            for _, row in bot5.iterrows():
                score = row["composite"]
                color = GREEN if score >= 75 else (ORANGE if score >= 50 else RED)
                st.markdown(f'<div style="display:flex; justify-content:space-between; align-items:center; padding:0.5rem 0.8rem; margin:0.3rem 0; background:#FFFFFF; border-left:4px solid {color}; border-radius:4px; border:1px solid #E8ECF0;">'
                            f'<span style="color:#1F2937; font-weight:600;">#{int(row["rank"])} {row["short_name"]}</span>'
                            f'<span style="color:{color}; font-weight:700; font-size:1.1rem;">{score:.0f}</span></div>', unsafe_allow_html=True)

    # Composite score chart
    st.markdown('<div class="section-title">Composite Score by Store</div>', unsafe_allow_html=True)
    sc_colors = [GREEN if v >= 75 else (ORANGE if v >= 50 else RED) for v in sc_df["composite"]]
    fig_sc = go.Figure(go.Bar(
        x=sc_df["short_name"], y=sc_df["composite"],
        marker_color=sc_colors,
        text=sc_df["composite"].apply(lambda x: f"{x:.0f}"),
        textposition="outside", textfont=dict(size=9, color="#374151"),
        hovertemplate="%{x}<br>Score: %{y:.1f}<extra></extra>",
    ))
    sc_layout = {**CHART_LAYOUT, "yaxis": dict(range=[0, 110], gridcolor=GRID_COLOR, fixedrange=True)}
    fig_sc.update_layout(**sc_layout, height=400, yaxis_title="Composite Score", xaxis_tickangle=-45)
    st.plotly_chart(fig_sc, use_container_width=True, key="sc_composite", config=CHART_CONFIG)

    # Stoplight table
    st.markdown('<div class="section-title">Scorecard Detail</div>', unsafe_allow_html=True)

    def stoplight(val, good, warn_fn="lower"):
        if pd.isna(val): return "—"
        if warn_fn == "lower":
            color = GREEN if val <= good else (ORANGE if val <= good * 1.3 else RED)
        else:
            color = GREEN if val >= good else (ORANGE if val >= good * 0.85 else RED)
        return f'<span style="color:{color}; font-weight:600;">{val:.1f}</span>'

    sc_tbl = sc_df[["rank", "short_name", "district"]].copy()
    sc_tbl["Composite"] = sc_df["composite"].apply(lambda x: f"{x:.0f}" if pd.notna(x) else "—")
    sc_tbl["SOS"] = sc_df["sos_min"].apply(fmt_sos) if "sos_min" in sc_df.columns else "—"
    sc_tbl["Pre-Bump"] = sc_df["pre_bump"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "—") if "pre_bump" in sc_df.columns else "—"
    sc_tbl["Adoption"] = sc_df["bone_in_adopt"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—") if "bone_in_adopt" in sc_df.columns else "—"
    sc_tbl["Waste"] = sc_df["waste"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "—") if "waste" in sc_df.columns else "—"
    sc_tbl["Labor %"] = sc_df["labor_pct"].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "—") if "labor_pct" in sc_df.columns else "—"
    sc_tbl["Dissat %"] = sc_df["dissat_pct"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—") if "dissat_pct" in sc_df.columns else "—"
    sc_tbl.columns = ["Rank", "Store", "District", "Score", "SOS (min)", "Pre-Bump %", "Adoption %", "Waste %", "Labor %", "Dissat %"]
    st.dataframe(sc_tbl, use_container_width=True, hide_index=True)

    # District scorecard
    if len(sc_df) > 5:
        st.markdown('<div class="section-title">District Average Scores</div>', unsafe_allow_html=True)
        dist_sc = sc_df.groupby("district")["composite"].mean().reset_index()
        dist_sc = dist_sc.sort_values("composite", ascending=False)
        dist_sc.columns = ["District", "Avg Score"]
        dist_colors = [GREEN if v >= 75 else (ORANGE if v >= 50 else RED) for v in dist_sc["Avg Score"]]
        fig_dsc = go.Figure(go.Bar(
            x=dist_sc["District"], y=dist_sc["Avg Score"],
            marker_color=dist_colors,
            text=dist_sc["Avg Score"].apply(lambda x: f"{x:.0f}"),
            textposition="outside", textfont=dict(size=11, color="#374151"),
            hovertemplate="%{x}<br>Score: %{y:.1f}<extra></extra>",
        ))
        dsc_layout = {**CHART_LAYOUT, "yaxis": dict(range=[0, 110], gridcolor=GRID_COLOR, fixedrange=True)}
        fig_dsc.update_layout(**dsc_layout, height=350, yaxis_title="Avg Composite Score")
        st.plotly_chart(fig_dsc, use_container_width=True, key="sc_district", config=CHART_CONFIG)

# ════════════════════════════════
# WATCH LIST
# ════════════════════════════════
elif selected_tab == "Watch List":

    # ── HEADER ──
    st.markdown(f"""
    <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.8rem;">
        <div>
            <h2 style="color:#1A3C34; font-weight:800; margin:0; font-size:1.6rem;">Watch List</h2>
            <p style="color:#6B7280; font-size:0.88rem; margin:0.2rem 0 0 0;">
                Auto-generated alerts from KDS, Labor, COGS &amp; QSC data &nbsp;·&nbsp; {len(DISTRICTS)} districts
            </p>
        </div>
        <div style="background:#DC2626; color:#FFFFFF; padding:0.5rem 1.2rem; border-radius:8px; font-weight:700; font-size:0.9rem; white-space:nowrap;">
            🚨 THRESHOLD ALERTS
        </div>
    </div>
    """, unsafe_allow_html=True)

    alerts = []

    # ── KDS Alerts (latest week from kds_dinner.csv) ──
    kds_wl_file = DATA_DIR / "kds_dinner.csv"
    if kds_wl_file.exists():
        kds_wl = pd.read_csv(kds_wl_file)
        kds_wl_periods = sorted(kds_wl["Period"].unique(), key=lambda x: (int(x[1]), int(x[3])))
        kds_wl_latest = kds_wl_periods[-1] if kds_wl_periods else ""
        kds_latest = kds_wl[kds_wl["Period"] == kds_wl_latest].copy()
        kds_latest["store_num"] = kds_latest["Store No"].astype(str)
        kds_latest["district"] = kds_latest["store_num"].map(STORE_TO_DISTRICT).fillna("Unassigned")

        # SOS > 10 min
        for _, r in kds_latest[kds_latest["SOS"] > 10].iterrows():
            sev = "Critical" if r["SOS"] >= 13 else "Warning"
            alerts.append({"store_num": r["store_num"], "Store": str(r["Store Name"])[:22], "District": r["district"],
                           "Metric": "SOS", "Value": fmt_sos(r['SOS']), "Threshold": "≤ 10:00", "Severity": sev, "Source": f"KDS {kds_wl_latest}"})

        # Adoption < 85%
        for _, r in kds_latest[kds_latest["Adoption %"].notna() & (kds_latest["Adoption %"] < 85)].iterrows():
            alerts.append({"store_num": r["store_num"], "Store": str(r["Store Name"])[:22], "District": r["district"],
                           "Metric": "Adoption", "Value": f"{r['Adoption %']:.1f}%", "Threshold": "≥ 85%", "Severity": "Critical", "Source": f"KDS {kds_wl_latest}"})

        # Make Ahead > 10%
        for _, r in kds_latest[kds_latest["Make Ahead %"].notna() & (kds_latest["Make Ahead %"] > 10)].iterrows():
            alerts.append({"store_num": r["store_num"], "Store": str(r["Store Name"])[:22], "District": r["district"],
                           "Metric": "Make Ahead", "Value": f"{r['Make Ahead %']:.1f}%", "Threshold": "≤ 10%", "Severity": "Critical", "Source": f"KDS {kds_wl_latest}"})

        # Pre-Bump > 1.5%
        for _, r in kds_latest[kds_latest["Pre-Bump %"].notna() & (kds_latest["Pre-Bump %"] > 1.5)].iterrows():
            alerts.append({"store_num": r["store_num"], "Store": str(r["Store Name"])[:22], "District": r["district"],
                           "Metric": "Pre-Bump", "Value": f"{r['Pre-Bump %']:.2f}%", "Threshold": "≤ 1.5%", "Severity": "Critical", "Source": f"KDS {kds_wl_latest}"})

        # Waste > 5%
        for _, r in kds_latest[kds_latest["Waste %"].notna() & (kds_latest["Waste %"] > 5)].iterrows():
            alerts.append({"store_num": r["store_num"], "Store": str(r["Store Name"])[:22], "District": r["district"],
                           "Metric": "Waste", "Value": f"{r['Waste %']:.2f}%", "Threshold": "≤ 5%", "Severity": "Critical", "Source": f"KDS {kds_wl_latest}"})

    # ── Labor Alerts ──
    forecast_path = DATA_DIR / "forecast.xlsm"
    if forecast_path.exists():
        @st.cache_data(ttl=300)
        def load_wl_labor():
            raw = pd.read_excel(forecast_path, sheet_name="Forecast_Data")
            return raw[raw["year"] == 2026].copy()
        wl_labor = load_wl_labor()
        if not wl_labor.empty:
            wl_labor["store_num"] = wl_labor["store"].apply(forecast_store_num)
            wl_labor["short_name"] = wl_labor["store"].apply(forecast_short_name)
            la = wl_labor.groupby(["store_num", "short_name"]).agg(
                actual_sales=("actual_sales", "sum"), forecast_sales=("forecast_sales", "sum"),
                actual_labor=("actual_labor", "sum"), schedule_labor=("schedule_labor", "sum"),
                ovt_hours=("ovt_hours", "sum"),
            ).reset_index()
            la["labor_pct"] = la["actual_labor"] / la["actual_sales"]
            la["labor_var"] = la["labor_pct"] - (la["schedule_labor"] / la["forecast_sales"])
            la["district"] = la["store_num"].map(STORE_TO_DISTRICT).fillna("Unassigned")

            for _, r in la[la["labor_pct"] > 0.18].iterrows():
                sev = "Critical" if r["labor_pct"] > 0.20 else "Warning"
                alerts.append({"store_num": r["store_num"], "Store": r["short_name"], "District": r["district"],
                               "Metric": "Labor %", "Value": f"{r['labor_pct']:.1%}", "Threshold": "≤ 18%", "Severity": sev, "Source": "Labor YTD"})

            for _, r in la[la["labor_var"] > 0.02].iterrows():
                alerts.append({"store_num": r["store_num"], "Store": r["short_name"], "District": r["district"],
                               "Metric": "Labor Variance", "Value": f"{r['labor_var']:+.2%}", "Threshold": "≤ +2%", "Severity": "Critical", "Source": "Labor YTD"})

            for _, r in la[la["ovt_hours"] > 25].iterrows():
                alerts.append({"store_num": r["store_num"], "Store": r["short_name"], "District": r["district"],
                               "Metric": "Overtime", "Value": f"{r['ovt_hours']:.0f} hrs", "Threshold": "≤ 25 hrs", "Severity": "Warning", "Source": "Labor YTD"})

    # ── COGS Alerts ──
    cogs_wl_file = DATA_DIR / "cogs_variance.csv"
    if cogs_wl_file.exists():
        cogs_wl = pd.read_csv(cogs_wl_file)
        cogs_latest_period = sorted(cogs_wl["Period"].unique())[-1] if len(cogs_wl) > 0 else ""
        cogs_latest = cogs_wl[cogs_wl["Period"] == cogs_latest_period].copy()
        for _, r in cogs_latest[cogs_latest["COGS Variance %"].notna() & (cogs_latest["COGS Variance %"] > 2)].iterrows():
            sev = "Critical" if r["COGS Variance %"] > 4 else "Warning"
            snum = str(r["Store No"]).lstrip("0")
            dist = STORE_TO_DISTRICT.get(snum, "Unassigned")
            alerts.append({"store_num": snum, "Store": str(r["Store Name"])[:22], "District": dist,
                           "Metric": "COGS Variance", "Value": f"{r['COGS Variance %']:.2f}%", "Threshold": "≤ 2%", "Severity": sev, "Source": f"COGS {cogs_latest_period}"})

    # ── QSC Eval Alerts ──
    eval_wl_file = DATA_DIR / "qsc_evals.csv"
    if eval_wl_file.exists():
        eval_wl = pd.read_csv(eval_wl_file)
        eval_wl["No Eval"] = eval_wl["No Eval"].astype(bool)
        eval_wl["Red Flag"] = eval_wl["Red Flag"].astype(bool)
        eval_periods = sorted(eval_wl["Period"].unique(), key=lambda x: (int(x[1]), int(x[3])))
        eval_latest = eval_periods[-1] if eval_periods else ""
        eval_latest_data = eval_wl[eval_wl["Period"] == eval_latest]

        for _, r in eval_latest_data[eval_latest_data["No Eval"]].iterrows():
            snum = str(int(r["Store No"]))
            dist = STORE_TO_DISTRICT.get(snum, "Unassigned")
            alerts.append({"store_num": snum, "Store": f"Store {snum}", "District": dist,
                           "Metric": "QSC Eval", "Value": "Missed", "Threshold": "Completed", "Severity": "Critical", "Source": f"QSC {eval_latest}"})

        for _, r in eval_latest_data[eval_latest_data["Red Flag"]].iterrows():
            snum = str(int(r["Store No"]))
            dist = STORE_TO_DISTRICT.get(snum, "Unassigned")
            reason = "score=0" if (pd.notna(r.get("Score")) and r["Score"] == 0) else "<1hr duration"
            alerts.append({"store_num": snum, "Store": f"Store {snum}", "District": dist,
                           "Metric": "QSC Red Flag", "Value": reason, "Threshold": "No flags", "Severity": "Warning", "Source": f"QSC {eval_latest}"})

    # ── Apply sidebar filters ──
    if alerts:
        alert_df = pd.DataFrame(alerts)
        if selected_store != "All Stores":
            sk_num = extract_store_number(selected_store)
            alert_df = alert_df[alert_df["store_num"] == sk_num]
        elif selected_district != "All Districts":
            d_nums = {s.split(" - ")[0].strip().lstrip("0") for s in DISTRICTS.get(selected_district, [])}
            alert_df = alert_df[alert_df["store_num"].isin(d_nums)]
    else:
        alert_df = pd.DataFrame()

    if alert_df.empty:
        st.success("✅ All stores within thresholds. No alerts.")
    else:
        # ── KPIs ──
        critical_count = (alert_df["Severity"] == "Critical").sum()
        warning_count = (alert_df["Severity"] == "Warning").sum()
        stores_flagged = alert_df["store_num"].nunique()
        districts_flagged = alert_df["District"].nunique()

        kpi_style = """<div style="background:#FFFFFF; border:1px solid #E2E8F0; border-radius:10px; padding:1rem; text-align:left; box-shadow:0 1px 3px rgba(0,0,0,0.04);">
            <div style="color:#6B7280; font-size:0.72rem; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">{label}</div>
            <div style="color:{color}; font-size:2rem; font-weight:800; margin:0.2rem 0;">{value}</div>
            <div style="color:#9CA3AF; font-size:0.78rem;">{sub}</div>
        </div>"""

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(kpi_style.format(label="CRITICAL ALERTS", value=critical_count, color="#DC2626",
                    sub="immediate attention"), unsafe_allow_html=True)
        c2.markdown(kpi_style.format(label="WARNINGS", value=warning_count, color="#D97706",
                    sub="monitor closely"), unsafe_allow_html=True)
        c3.markdown(kpi_style.format(label="STORES FLAGGED", value=f"{stores_flagged}/{len(STORE_TO_DISTRICT)}", color="#1F2937",
                    sub=f"{stores_flagged/len(STORE_TO_DISTRICT)*100:.0f}% of fleet"), unsafe_allow_html=True)
        c4.markdown(kpi_style.format(label="DISTRICTS AFFECTED", value=f"{districts_flagged}/{len(DISTRICTS)}", color="#1F2937",
                    sub="with alerts"), unsafe_allow_html=True)

        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

        # ── Takeaways ──
        st.markdown('<div style="font-weight:700; color:#1A3C34; font-size:1.05rem; margin:0.5rem 0 0.5rem 0;">Alert Summary</div>', unsafe_allow_html=True)
        takeaway_style = '<div style="border-left:4px solid {color}; padding:0.5rem 1rem; margin:0.4rem 0; background:#FAFBFC; border-radius:0 6px 6px 0;">{text}</div>'

        # Most common alert types
        metric_counts = alert_df["Metric"].value_counts()
        top_metric = metric_counts.index[0]
        st.markdown(takeaway_style.format(color="#1A3C34",
            text=f'<b>{len(alert_df)} total alerts</b> across {stores_flagged} stores — <b>{critical_count} critical</b>, {warning_count} warnings.'),
            unsafe_allow_html=True)

        st.markdown(takeaway_style.format(color="#DC2626",
            text=f'<b>Most common issue:</b> {top_metric} ({metric_counts.iloc[0]} alerts). ' +
                 ", ".join(f"{m}: {c}" for m, c in metric_counts.items() if m != top_metric)),
            unsafe_allow_html=True)

        # Worst district
        dist_alert_counts = alert_df.groupby("District").size().sort_values(ascending=False)
        worst_dist = dist_alert_counts.index[0]
        worst_dist_crit = (alert_df[alert_df["District"] == worst_dist]["Severity"] == "Critical").sum()
        st.markdown(takeaway_style.format(color="#D97706",
            text=f'<b>Most alerts:</b> {worst_dist} — {dist_alert_counts.iloc[0]} alerts ({worst_dist_crit} critical).'),
            unsafe_allow_html=True)

        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

        # ── All stores by district ──
        st.markdown('<div class="section-title" style="font-size:1.05rem;">All Stores by District</div>', unsafe_allow_html=True)

        # Build alert lookup: store_num -> list of alerts
        alert_by_store = {}
        for _, a in alert_df.iterrows():
            alert_by_store.setdefault(a["store_num"], []).append(a)

        # Determine which districts to show
        if selected_district != "All Districts":
            districts_to_show = {selected_district: DISTRICTS[selected_district]}
        else:
            districts_to_show = DISTRICTS

        for district, stores in sorted(districts_to_show.items()):
            # Count alerts for this district
            d_store_nums = [s.split(" - ")[0].strip().lstrip("0") for s in stores]
            d_alerts_list = [a for sn in d_store_nums for a in alert_by_store.get(sn, [])]
            d_crit = sum(1 for a in d_alerts_list if a["Severity"] == "Critical")
            d_warn = sum(1 for a in d_alerts_list if a["Severity"] == "Warning")
            d_flagged = sum(1 for sn in d_store_nums if sn in alert_by_store)
            d_clean = len(d_store_nums) - d_flagged

            badge_html = ""
            if d_crit > 0:
                badge_html += f'<span style="background:#DC2626; color:white; font-size:0.7rem; font-weight:700; padding:0.15rem 0.5rem; border-radius:10px; margin-left:6px;">{d_crit} critical</span>'
            if d_warn > 0:
                badge_html += f'<span style="background:#D97706; color:white; font-size:0.7rem; font-weight:700; padding:0.15rem 0.5rem; border-radius:10px; margin-left:4px;">{d_warn} warning</span>'
            if d_clean > 0:
                badge_html += f'<span style="background:#059669; color:white; font-size:0.7rem; font-weight:700; padding:0.15rem 0.5rem; border-radius:10px; margin-left:4px;">{d_clean} clear</span>'

            st.markdown(f"""
            <div style="background:#1A3C34; color:#FFFFFF; padding:0.5rem 1rem; border-radius:6px 6px 0 0; margin-top:1rem;
                        display:flex; justify-content:space-between; align-items:center;">
                <span style="font-weight:700; font-size:0.95rem;">{district}{badge_html}</span>
                <span style="font-size:0.82rem;">{len(d_store_nums)} stores &nbsp;|&nbsp; {len(d_alerts_list)} alerts</span>
            </div>
            """, unsafe_allow_html=True)

            cards_html = ""
            for s in stores:
                snum = s.split(" - ")[0].strip().lstrip("0")
                sname = s.split("-", 2)[2].strip()[:25] if len(s.split("-", 2)) >= 3 else s[:25]
                s_alerts = alert_by_store.get(snum, [])

                if s_alerts:
                    has_crit = any(a["Severity"] == "Critical" for a in s_alerts)
                    border = "#DC2626" if has_crit else "#D97706"
                    bg = "#FEF2F2" if has_crit else "#FFFBEB"
                    status_icon = f'<span style="color:#DC2626; font-weight:700;">⚠ {len(s_alerts)}</span>' if has_crit else f'<span style="color:#D97706; font-weight:700;">⚠ {len(s_alerts)}</span>'

                    issues = ""
                    for a in s_alerts:
                        dot_c = "#DC2626" if a["Severity"] == "Critical" else "#D97706"
                        issues += f'<div style="padding:0.1rem 0; font-size:0.8rem; color:#374151;"><span style="color:{dot_c};">●</span> <b>{a["Metric"]}</b>: {a["Value"]} <span style="color:#9CA3AF;">({a["Source"]})</span></div>'

                    cards_html += f"""<div style="border-left:4px solid {border}; padding:0.5rem 0.8rem; border-bottom:1px solid #F1F5F9; background:{bg};">
                        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:0.2rem;">
                            <span style="color:#1F2937; font-weight:700; font-size:0.88rem;">#{snum} {sname}</span>
                            {status_icon}
                        </div>
                        {issues}
                    </div>"""
                else:
                    cards_html += f"""<div style="border-left:4px solid #059669; padding:0.5rem 0.8rem; border-bottom:1px solid #F1F5F9; background:#FFFFFF;">
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span style="color:#1F2937; font-weight:600; font-size:0.88rem;">#{snum} {sname}</span>
                            <span style="color:#059669; font-weight:700; font-size:0.82rem;">✓ Clear</span>
                        </div>
                    </div>"""

            st.markdown(f'<div style="border:1px solid #E2E8F0; border-radius:0 0 6px 6px; overflow:hidden; margin-bottom:0.5rem;">{cards_html}</div>', unsafe_allow_html=True)

        # ── Threshold Reference ──
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="font-size:1.05rem;">Threshold Reference</div>', unsafe_allow_html=True)
        ref_html = """
        <table style="width:100%; border-collapse:collapse; border:1px solid #E2E8F0; border-radius:6px; overflow:hidden;">
            <thead><tr style="background:#F1F5F9;">
                <th style="padding:0.5rem 0.8rem; text-align:left; font-size:0.75rem; color:#6B7280; font-weight:600; text-transform:uppercase; border-bottom:2px solid #E2E8F0;">Metric</th>
                <th style="padding:0.5rem 0.8rem; text-align:center; font-size:0.75rem; color:#6B7280; font-weight:600; text-transform:uppercase; border-bottom:2px solid #E2E8F0;">Threshold</th>
                <th style="padding:0.5rem 0.8rem; text-align:center; font-size:0.75rem; color:#6B7280; font-weight:600; text-transform:uppercase; border-bottom:2px solid #E2E8F0;">Severity</th>
                <th style="padding:0.5rem 0.8rem; text-align:left; font-size:0.75rem; color:#6B7280; font-weight:600; text-transform:uppercase; border-bottom:2px solid #E2E8F0;">Source</th>
            </tr></thead><tbody>"""
        thresholds = [
            ("SOS", "> 10 min (≥13 critical)", "Warning / Critical", "KDS Dinner"),
            ("Adoption %", "< 85%", "Critical", "KDS Dinner"),
            ("Make Ahead %", "> 10%", "Critical", "KDS Dinner"),
            ("Pre-Bump %", "> 1.5%", "Critical", "KDS Dinner"),
            ("Waste %", "> 5%", "Critical", "KDS Dinner"),
            ("Labor %", "> 18% (>20% critical)", "Warning / Critical", "Labor Forecast"),
            ("Labor Variance", "> +2%", "Critical", "Labor Forecast"),
            ("Overtime", "> 25 hrs", "Warning", "Labor Forecast"),
            ("COGS Variance", "> 2% (>4% critical)", "Warning / Critical", "COGS Data"),
            ("QSC Eval", "Missed evaluation", "Critical", "QSC Evals"),
            ("QSC Red Flag", "Score=0 or <1hr", "Warning", "QSC Evals"),
        ]
        for i, (metric, thresh, sev, source) in enumerate(thresholds):
            bg = "#FFFFFF" if i % 2 == 0 else "#F9FAFB"
            sev_color = "#DC2626" if "Critical" in sev else "#D97706"
            ref_html += f'<tr style="background:{bg};"><td style="padding:0.4rem 0.8rem; border-bottom:1px solid #F1F5F9; font-weight:600; color:#1F2937;">{metric}</td><td style="padding:0.4rem 0.8rem; border-bottom:1px solid #F1F5F9; text-align:center; color:#374151;">{thresh}</td><td style="padding:0.4rem 0.8rem; border-bottom:1px solid #F1F5F9; text-align:center; color:{sev_color}; font-weight:600;">{sev}</td><td style="padding:0.4rem 0.8rem; border-bottom:1px solid #F1F5F9; color:#6B7280;">{source}</td></tr>'
        ref_html += "</tbody></table>"
        st.markdown(ref_html, unsafe_allow_html=True)

# ════════════════════════════════
# WING WORM
# ════════════════════════════════
elif selected_tab == "Wing Worm":
    st.markdown('<div class="section-title">Wing Worm</div>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Guide the worm to eat boneless wings! Use arrow keys or WASD to move.</p>', unsafe_allow_html=True)

    import streamlit.components.v1 as components
    game_html = """
    <div id="game-wrapper" style="display:flex; flex-direction:column; align-items:center;">
        <div style="display:flex; justify-content:space-between; width:400px; margin-bottom:8px;">
            <span id="score" style="color:#1F2937; font-weight:700; font-size:1rem;">Score: 0</span>
            <span id="high-score" style="color:#059669; font-weight:700; font-size:1rem;">Best: 0</span>
        </div>
        <canvas id="game" width="400" height="400" style="border:2px solid #0D5C3F; border-radius:8px; background:#0A4A32;"></canvas>
        <div id="game-over" style="display:none; margin-top:12px; text-align:center;">
            <p style="color:#DC2626; font-weight:700; font-size:1.2rem; margin:0;">Game Over!</p>
            <p style="color:#6B7280; font-size:0.85rem; margin:4px 0;">Press Space or tap to restart</p>
        </div>
        <div id="start-msg" style="margin-top:12px; text-align:center;">
            <p style="color:#059669; font-weight:600; font-size:0.95rem; margin:0;">Press any arrow key or tap to start!</p>
        </div>
    </div>
    <script>
    (function() {
        const canvas = document.getElementById('game');
        const ctx = canvas.getContext('2d');
        const GRID = 20;
        const COLS = canvas.width / GRID;
        const ROWS = canvas.height / GRID;

        let snake, dir, nextDir, food, score, highScore, gameOver, started, interval, lastTime, accumulator;

        highScore = parseInt(localStorage.getItem('wingworm_high') || '0');
        document.getElementById('high-score').textContent = 'Best: ' + highScore;

        function init() {
            snake = [{x:10, y:10}, {x:9, y:10}, {x:8, y:10}];
            dir = {x:1, y:0};
            nextDir = {x:1, y:0};
            score = 0;
            gameOver = false;
            started = false;
            interval = 120;
            lastTime = 0;
            accumulator = 0;
            placeFood();
            updateScore();
            document.getElementById('game-over').style.display = 'none';
            document.getElementById('start-msg').style.display = 'block';
            draw();
        }

        function placeFood() {
            do {
                food = {x: Math.floor(Math.random() * COLS), y: Math.floor(Math.random() * ROWS)};
            } while (snake.some(s => s.x === food.x && s.y === food.y));
        }

        function updateScore() {
            document.getElementById('score').textContent = 'Score: ' + score;
            if (score > highScore) {
                highScore = score;
                localStorage.setItem('wingworm_high', highScore);
                document.getElementById('high-score').textContent = 'Best: ' + highScore;
            }
        }

        function drawRoundRect(x, y, w, h, r, color) {
            ctx.fillStyle = color;
            ctx.beginPath();
            ctx.moveTo(x+r, y);
            ctx.arcTo(x+w, y, x+w, y+h, r);
            ctx.arcTo(x+w, y+h, x, y+h, r);
            ctx.arcTo(x, y+h, x, y, r);
            ctx.arcTo(x, y, x+w, y, r);
            ctx.closePath();
            ctx.fill();
        }

        function draw() {
            // Dark green gradient background
            const grad = ctx.createLinearGradient(0, 0, 0, canvas.height);
            grad.addColorStop(0, '#0A4A32');
            grad.addColorStop(1, '#0D5C3F');
            ctx.fillStyle = grad;
            ctx.fillRect(0, 0, canvas.width, canvas.height);

            // Subtle grid
            ctx.strokeStyle = 'rgba(255,255,255,0.06)';
            ctx.lineWidth = 0.5;
            for (let i = 0; i < COLS; i++) {
                ctx.beginPath(); ctx.moveTo(i*GRID, 0); ctx.lineTo(i*GRID, canvas.height); ctx.stroke();
            }
            for (let i = 0; i < ROWS; i++) {
                ctx.beginPath(); ctx.moveTo(0, i*GRID); ctx.lineTo(canvas.width, i*GRID); ctx.stroke();
            }

            // Food (boneless wing) with glow
            const fx = food.x * GRID, fy = food.y * GRID;
            ctx.shadowColor = '#F59E0B';
            ctx.shadowBlur = 10;
            ctx.fillStyle = '#D97706';
            ctx.beginPath();
            ctx.ellipse(fx + GRID/2, fy + GRID/2, GRID/2.2, GRID/2.8, 0, 0, Math.PI*2);
            ctx.fill();
            ctx.shadowBlur = 0;
            ctx.fillStyle = '#F59E0B';
            ctx.beginPath();
            ctx.ellipse(fx + GRID/2, fy + GRID/2.5, GRID/3.5, GRID/4, 0, 0, Math.PI*2);
            ctx.fill();

            // Snake with glow on head
            snake.forEach((seg, i) => {
                if (i === 0) {
                    ctx.shadowColor = '#34D399';
                    ctx.shadowBlur = 8;
                }
                const color = i === 0 ? '#10B981' : (i % 2 === 0 ? '#059669' : '#0D9488');
                drawRoundRect(seg.x * GRID + 1, seg.y * GRID + 1, GRID - 2, GRID - 2, 4, color);
                if (i === 0) ctx.shadowBlur = 0;
            });

            // Eyes on head
            const head = snake[0];
            const eyeSize = 3;
            ctx.fillStyle = '#FFFFFF';
            if (dir.x === 1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+15, head.y*GRID+6, eyeSize, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+15, head.y*GRID+14, eyeSize, 0, Math.PI*2); ctx.fill();
            } else if (dir.x === -1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+5, head.y*GRID+6, eyeSize, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+5, head.y*GRID+14, eyeSize, 0, Math.PI*2); ctx.fill();
            } else if (dir.y === -1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+6, head.y*GRID+5, eyeSize, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+14, head.y*GRID+5, eyeSize, 0, Math.PI*2); ctx.fill();
            } else {
                ctx.beginPath(); ctx.arc(head.x*GRID+6, head.y*GRID+15, eyeSize, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+14, head.y*GRID+15, eyeSize, 0, Math.PI*2); ctx.fill();
            }
            ctx.fillStyle = '#0A4A32';
            if (dir.x === 1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+16, head.y*GRID+6, 1.5, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+16, head.y*GRID+14, 1.5, 0, Math.PI*2); ctx.fill();
            } else if (dir.x === -1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+4, head.y*GRID+6, 1.5, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+4, head.y*GRID+14, 1.5, 0, Math.PI*2); ctx.fill();
            } else if (dir.y === -1) {
                ctx.beginPath(); ctx.arc(head.x*GRID+6, head.y*GRID+4, 1.5, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+14, head.y*GRID+4, 1.5, 0, Math.PI*2); ctx.fill();
            } else {
                ctx.beginPath(); ctx.arc(head.x*GRID+6, head.y*GRID+16, 1.5, 0, Math.PI*2); ctx.fill();
                ctx.beginPath(); ctx.arc(head.x*GRID+14, head.y*GRID+16, 1.5, 0, Math.PI*2); ctx.fill();
            }
        }

        function step() {
            if (gameOver) return;
            dir = nextDir;
            const head = {x: snake[0].x + dir.x, y: snake[0].y + dir.y};

            if (head.x < 0 || head.x >= COLS || head.y < 0 || head.y >= ROWS) { die(); return; }
            if (snake.some(s => s.x === head.x && s.y === head.y)) { die(); return; }

            snake.unshift(head);
            if (head.x === food.x && head.y === food.y) {
                score++;
                updateScore();
                placeFood();
                if (interval > 60) interval -= 2;
            } else {
                snake.pop();
            }
            draw();
        }

        function gameLoop(timestamp) {
            if (!started || gameOver) return;
            if (!lastTime) lastTime = timestamp;
            const delta = timestamp - lastTime;
            lastTime = timestamp;
            accumulator += delta;
            while (accumulator >= interval) {
                step();
                accumulator -= interval;
                if (gameOver) break;
            }
            requestAnimationFrame(gameLoop);
        }

        function die() {
            gameOver = true;
            document.getElementById('game-over').style.display = 'block';
            draw();
            ctx.fillStyle = 'rgba(0,0,0,0.5)';
            ctx.fillRect(0, 0, canvas.width, canvas.height);
            ctx.fillStyle = '#FFFFFF';
            ctx.font = 'bold 28px sans-serif';
            ctx.textAlign = 'center';
            ctx.fillText('Game Over!', canvas.width/2, canvas.height/2 - 10);
            ctx.font = '16px sans-serif';
            ctx.fillText('Score: ' + score + '  |  Wings eaten!', canvas.width/2, canvas.height/2 + 20);
        }

        function startGame() {
            if (!started) {
                started = true;
                lastTime = 0;
                accumulator = 0;
                document.getElementById('start-msg').style.display = 'none';
                requestAnimationFrame(gameLoop);
            }
        }

        document.addEventListener('keydown', function(e) {
            if (['ArrowUp','ArrowDown','ArrowLeft','ArrowRight','w','a','s','d',' '].includes(e.key)) {
                e.preventDefault();
            }
            if (gameOver && (e.key === ' ' || e.key === 'Enter')) { init(); return; }
            if (e.key === 'ArrowUp' || e.key === 'w') { if (dir.y !== 1) { nextDir = {x:0,y:-1}; startGame(); } }
            else if (e.key === 'ArrowDown' || e.key === 's') { if (dir.y !== -1) { nextDir = {x:0,y:1}; startGame(); } }
            else if (e.key === 'ArrowLeft' || e.key === 'a') { if (dir.x !== 1) { nextDir = {x:-1,y:0}; startGame(); } }
            else if (e.key === 'ArrowRight' || e.key === 'd') { if (dir.x !== -1) { nextDir = {x:1,y:0}; startGame(); } }
        });

        let touchStart = null;
        canvas.addEventListener('touchstart', function(e) {
            e.preventDefault();
            if (gameOver) { init(); return; }
            touchStart = {x: e.touches[0].clientX, y: e.touches[0].clientY};
            startGame();
        });
        canvas.addEventListener('touchmove', function(e) {
            e.preventDefault();
            if (!touchStart) return;
            const dx = e.touches[0].clientX - touchStart.x;
            const dy = e.touches[0].clientY - touchStart.y;
            if (Math.abs(dx) > Math.abs(dy)) {
                if (dx > 20 && dir.x !== -1) nextDir = {x:1, y:0};
                else if (dx < -20 && dir.x !== 1) nextDir = {x:-1, y:0};
            } else {
                if (dy > 20 && dir.y !== -1) nextDir = {x:0, y:1};
                else if (dy < -20 && dir.y !== 1) nextDir = {x:0, y:-1};
            }
            touchStart = {x: e.touches[0].clientX, y: e.touches[0].clientY};
        });

        canvas.addEventListener('click', function() {
            if (gameOver) init();
            else startGame();
        });

        init();
    })();
    </script>
    """
    components.html(game_html, height=500)

st.markdown("---")
st.markdown('<p style="color:#999999; font-size:0.75rem; text-align:center;">FL Wingmen Dashboard &nbsp;|&nbsp; Smart Kitchen Performance &amp; Forecast Data</p>', unsafe_allow_html=True)
